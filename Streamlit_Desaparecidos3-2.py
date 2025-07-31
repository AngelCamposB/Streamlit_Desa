# -------------------------
# Aplicación de Monitoreo de Personas Desaparecidas
# -------------------------
import re
import os
import time
import signal
import difflib
import threading
import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import matplotlib.pyplot as plt
import plotly.graph_objects as go

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from datetime import datetime
from statsmodels.tsa.arima.model import ARIMA

# -------------------------
# Carga y limpieza de datos
# -------------------------
@st.cache_data
def load_data(uploaded_file=None, default_filepath=None, columns_to_use=None):
    """
    Carga datos desde un archivo Excel, con prioridad al archivo subido.
    
    Args:
        uploaded_file: Archivo subido por el usuario
        default_filepath: Ruta del archivo por defecto
    
    Returns:
        DataFrame con los datos, o None si no se puede cargar
    """
    try:
        # Primero intenta cargar el archivo subido
        if uploaded_file is not None:
            # Si necesitamos solo algunas columnas y el formato es Excel
            if columns_to_use:
                # Primero leemos solo la primera fila para obtener las columnas
                headers_df = pd.read_excel(uploaded_file, nrows=0)
                # Verificamos que las columnas solicitadas existan
                available_columns = [col for col in columns_to_use if col in headers_df.columns]
                # Si faltan columnas, mostramos una advertencia
                if len(available_columns) < len(columns_to_use):
                    missing = set(columns_to_use) - set(available_columns)
                    st.warning(f"⚠️ Algunas columnas solicitadas no existen en el archivo: {missing}")
                # Leemos solo las columnas necesarias
                return pd.read_excel(uploaded_file, usecols=available_columns), uploaded_file.name
            else:
                return pd.read_excel(uploaded_file), uploaded_file.name
        
        # Si no hay archivo subido, intenta cargar el archivo por defecto
        if default_filepath is not None:
            if columns_to_use:
                # Primero leemos solo la primera fila para obtener las columnas
                headers_df = pd.read_excel(default_filepath, nrows=0)
                # Verificamos que las columnas solicitadas existan
                available_columns = [col for col in columns_to_use if col in headers_df.columns]
                # Si faltan columnas, mostramos una advertencia
                if len(available_columns) < len(columns_to_use):
                    missing = set(columns_to_use) - set(available_columns)
                    st.warning(f"⚠️ Algunas columnas solicitadas no existen en el archivo: {missing}")
                # Leemos solo las columnas necesarias
                return pd.read_excel(default_filepath, usecols=available_columns), os.path.basename(default_filepath)
            else:
                return pd.read_excel(default_filepath), os.path.basename(default_filepath)
        
        # Si no hay ningún archivo
        return None, None
    
    except FileNotFoundError:
        st.warning(f"⚠️ No se encontró el archivo Excel en: {default_filepath}")
        return None, None
    except PermissionError:
        st.error(f"🚫 Error de permisos al intentar acceder al archivo: {default_filepath}")
        return None, None
    except Exception as e:
        st.error(f"❌ Error al cargar el archivo Excel: {e}")
        return None, None

#---- Funciones para normalizar DD_ULTLUGAR ----#
def quitar_acentos(texto):
    """Reemplaza caracteres acentuados por sus equivalentes sin acento, conservando 'ñ' y 'Ñ'."""
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u',
        'ã': 'a', 'õ': 'o',
        'ä': 'a', 'ë': 'e', 'ï': 'i', 'ö': 'o', 'ü': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O', 'Û': 'U',
        'Ã': 'A', 'Õ': 'O',
        'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U'
    }
    for letra_acentuada, letra_sin_acento in reemplazos.items():
        texto = texto.replace(letra_acentuada, letra_sin_acento)
    return texto

def similarity(a: str, b: str) -> float:
    """
    Calcula la similitud entre dos cadenas usando SequenceMatcher.
    
    Args:
        a (str): Primera cadena.
        b (str): Segunda cadena.

    Returns:
        float: Ratio de similitud entre 0 y 1.
    """
    return difflib.SequenceMatcher(None, a, b).ratio()

def normalize_ultimo_lugar(texto: str, threshold: float = 0.75) -> str:
    """
    Normaliza cada palabra dentro de DD_ULTLUGAR corrigiendo errores de digitación
    mediante comparación con valores canónicos usando SequenceMatcher.
    
    Args:
        texto (str): Frase original con posibles errores tipográficos.
        valores_canonicos (list): Lista de palabras correctas de referencia.
        threshold (float): Umbral de similitud para considerar una corrección.
    
    Returns:
        str: Valor normalizado.
    """
    if pd.isna(texto) or texto.strip() == "":
        return np.nan

    # Preprocesar: eliminar espacios extra, pasar a minúsculas, quitar acentos y dividir en palabras
    palabras = quitar_acentos(texto.strip().lower()).split()
    
    # Lista de valores canónicos esperados para DD_ULTLUGAR
    valores_canonicos = [
        "ado",
        "aeropuerto",
        "agencia 29",
        "albergue",
        "alrededores de hunucma",
        "alta mar",
        "anexo",
        "bar",
        "caimede",
        "cantina",
        "casa de sus padres",
        "centro de rehabilitacion",
        "cereso",
        "cine",
        "domicilio",
        "convento",
        "detenido",
        "madre",
        "padre",
        "escuela",
        "transporte",
        "ciudad",
        "gimnasio",
        "guarderia",
        "hospital",
        "iglesia",
        "infonavit",
        "institución",
        "lugar publico"
    ]
    
    # Corregir cada palabra usando la lista de valores canónicos
    palabras_corregidas = []
    for palabra in palabras:
        mejor_coincidencia = max(valores_canonicos, key=lambda x: similarity(palabra, x))
        similitud_max = similarity(palabra, mejor_coincidencia)

        # Si la similitud es mayor al umbral, se usa la palabra corregida
        if similitud_max >= threshold:
            palabras_corregidas.append(mejor_coincidencia)
        else:
            palabras_corregidas.append(palabra)  # Mantener la palabra original si no hay coincidencia

    # Unir palabras corregidas en una frase con capitalización adecuada
    return " ".join(palabras_corregidas).capitalize()

def mapear_conceptos(texto: str) -> str:
    """
    Mapea frases similares en concepto a una forma estandarizada.

    Args:
        texto (str): Texto corregido.

    Returns:
        str: Texto con los conceptos unificados.
    """
    if pd.isna(texto) or texto.strip() == "":
        return np.nan
    
    # Diccionario de equivalencias
    mapeo = {
        r"(en)?( )?alta( )?mar": "En alta mar",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)?( )?(su)?( )?familiar": "Domicilio de su familiar",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de)?( )?(sus)?( )?(padres|papas)": "Domicilio de sus padres",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de)?( )?(sus)?( )?hermanos": "Domicilio de sus hermanos",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de)?( )?(sus)?( )?amigos": "Domicilio de sus amigos",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)?( )?(su)?( )?(mama|madre)": "Domicilio de su madre",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)?( )?(su)?( )?(papa|padre)": "Domicilio de su padre",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)?( )?(su)?( )?(esposo|marido)": "Domicilio de su esposo",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)?( )?(su)?( )?esposa": "Domicilio de su esposa",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)?( )?(su)?( )?(conyuge|conyugal)": "Domicilio de su conyuge",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)?( )?(su)?( )?hermano": "Domicilio de su hermano",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?hermana": "Domicilio de su hermana",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?hijo": "Domicilio de su hijo",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?hija": "Domicilio de su hija",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?abuelo": "Domicilio de su abuelo",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?abuela": "Domicilio de su abuela",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?tio": "Domicilio de su tio",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?tia": "Domicilio de su tia",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?primo": "Domicilio de su primo",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?prima": "Domicilio de su prima",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?amigo": "Domicilio de su amigo",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?amiga": "Domicilio de su amiga",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|del)( )?(su)?( )?cuñado": "Domicilio de su cuñado",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(de|de la)( )?(su)?( )?cuñada": "Domicilio de su cuñada",

        r"(en)?( )?(el)?( )?hoh": "Hospital OHoran",
        r"(en)?( )?(el)?( )?(hospital)?( )?(psi|psiq|psiquiatrico)": "Hospital Psiquiatrico",
        r"(en)?( )?(el)?( )?(lugar|centro)?(de)?( )?(trabajo|laboral)": "Trabajo",
        r"(en)?( )?(el)?( )?(lugar|centro)?(de)?( )?trabajo": "Trabajo",    #Por alguna extraña razon, es necesario para homologar.
        r"(en)?( )?(la)?( )?escuela": "Escuela",
        #r"\bs/\b": "s/d",
        r"\blugar publico\b": "Via publica",
        r"(en)?( )?(la)?( )?via( )?publica": "Via publica",
        r"(en)?( )?(su)?( )?(domicilio|casa)( )?(particular|propio)?": "Domicilio",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(que)?( )?renta( )?(el|la)?( )?(desaparecido|desaparecida)?": "Domicilio que renta",
        r"(en)?( )?(el|la)?( )?(domicilio|casa)( )?(que)?( )?rentaba( )?(el|la)?( )?(desaparecido|desaparecida)?": "Domicilio que rentaba",
    }

    # Aplicar reemplazos con expresiones regulares
    for patron, reemplazo in mapeo.items():
        texto = re.sub(patron, reemplazo, texto, flags=re.IGNORECASE)

    return texto.capitalize()
# ------------------------------------------------------ #
def normalize_ocupaciones(texto: str, threshold: float = 0.80) -> str:
    """
    Normaliza cada palabra dentro de DD_ULTLUGAR corrigiendo errores de digitación
    mediante comparación con valores canónicos usando SequenceMatcher.
    
    Args:
        texto (str): Frase original con posibles errores tipográficos.
        valores_canonicos (list): Lista de palabras correctas de referencia.
        threshold (float): Umbral de similitud para considerar una corrección.
    
    Returns:
        str: Valor normalizado.
    """
    if pd.isna(texto) or texto.strip() == "":
        return np.nan
    
    # Preprocesar: eliminar espacios extra, pasar a minúsculas, quitar acentos y dividir en palabras
    palabras = quitar_acentos(texto.strip().lower()).split()
    
    # Lista de valores canónicos esperados para DD_ULTLUGAR
    valores_canonicos = [
        "ayudante",
        "albañil",
        "bailarina",
        "bodeguero",
        "desempleado",
        #"desempleada",
        "cheff",
        "cocinero",
        "empleado",
        #"empleada",
        "enfermero",
        "enfermera",
        "estudiante",
        "pensionado",
        "comerciante",
        "cereso",
        "cine",
        "domicilio",
        "convento",
        "detenido",
        "madre",
        "padre",
        "escuela",
        "lugar público"
    ]
    
    # Corregir cada palabra usando la lista de valores canónicos
    palabras_corregidas = []
    for palabra in palabras:
        mejor_coincidencia = max(valores_canonicos, key=lambda x: similarity(palabra, x))
        similitud_max = similarity(palabra, mejor_coincidencia)

        # Si la similitud es mayor al umbral, se usa la palabra corregida
        if similitud_max >= threshold:
            palabras_corregidas.append(mejor_coincidencia)
        else:
            palabras_corregidas.append(palabra)  # Mantener la palabra original si no hay coincidencia

    # Unir palabras corregidas en una frase con capitalización adecuada
    return " ".join(palabras_corregidas).capitalize()
# ------------------------------------------------------ #

def convertir_a_fecha(valor):
            if pd.isna(valor) or valor == "": #verifica si el valor es nulo o vacio
                return np.nan #Retorna Not a Time, para que luego se pueda reemplazar con el valor de la columna CI_FECDEN.
            elif isinstance(valor, int):
                return pd.to_datetime(f'{valor}-01-01')
            elif isinstance(valor, str) and valor.isdigit() and len(valor) == 4:
                return pd.to_datetime(f'{valor}-01-01')
            else:
                try:
                    return pd.to_datetime(valor)
                except ValueError:
                    return pd.NaT
                
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Realiza la limpieza básica de los datos, considerando solo las columnas disponibles.
    Versión optimizada que procesa eficientemente las columnas presentes.
    """
    # Siempre trabajamos con una copia para evitar modificar el original
    df = df.copy()
    
    # Normalizar nombres de columnas
    df.columns = df.columns.str.strip()
    
    # Verificar si las columnas son numéricas (caso donde Excel carga con números)
    if list(df.columns) == list(range(len(df.columns))):
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)
    
    # Definir las columnas a procesar según su tipo
    columnas_numericas = ["PD_EDAD"]
    columnas_fecha_custom = ["CI_FECDEN", "DD_FECDESAP"]  # Fechas que usan convertir_a_fecha
    columnas_fecha_std = ["DL_FECLOC", "DD_FECPERCA"]     # Fechas que usan pd.to_datetime
    columnas_texto = ["PD_SEXO", "DD_MPIO", "DD_ESTADO", "CI_CARPEINV", "PD_OCUPA", "PD_NOMBRECOMPLETOVICFORMULA"]
    
    # --- PROCESAMIENTO DE COLUMNAS NUMÉRICAS ---
    for col in [c for c in columnas_numericas if c in df.columns]:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # --- PROCESAMIENTO DE COLUMNAS DE FECHA ---
    # Columnas de fecha que usan convertir_a_fecha
    for col in [c for c in columnas_fecha_custom if c in df.columns]:
        df[col] = df[col].apply(convertir_a_fecha)
    
    # Columnas de fecha que usan pd.to_datetime
    for col in [c for c in columnas_fecha_std if c in df.columns]:
        df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Llenar fechas de desaparición faltantes si ambas columnas están disponibles
    if "DD_FECDESAP" in df.columns and "CI_FECDEN" in df.columns:
        df['DD_FECDESAP'] = df['DD_FECDESAP'].fillna(df['CI_FECDEN'])
    
    # --- PROCESAMIENTO DE COLUMNAS DE TEXTO ---
    # Procesar PD_SEXO si está disponible (caso especial)
    if "PD_SEXO" in df.columns:
        mapeo_sexo = {
            'M': 'HOMBRE', 'MASCULINO': 'HOMBRE',
            'F': 'MUJER', 'FEMENINO': 'MUJER',
        }
        df["PD_SEXO"] = df["PD_SEXO"].str.strip().str.upper().map(mapeo_sexo).fillna(df["PD_SEXO"].str.strip().str.upper())
    
    # Procesar otras columnas de texto disponibles
    texto_cols = [c for c in columnas_texto if c in df.columns and c != "PD_SEXO"]
    for col in texto_cols:
        df[col] = df[col].astype(str).str.strip().str.upper()
    
    # --- PROCESAMIENTO ESPECIALIZADO ---
    # Normalización de DD_ULTLUGAR si está disponible
    if "DD_ULTLUGAR" in df.columns:
        df["DD_ULTLUGAR"] = df["DD_ULTLUGAR"].apply(normalize_ultimo_lugar)
        df["DD_ULTLUGAR"] = df["DD_ULTLUGAR"].apply(mapear_conceptos)
    
    # Procesar PD_OCUPA si está disponible
    if "PD_OCUPA" in df.columns:
        # Reemplazar cadenas vacías con None para que Plotly Express las omita
        df.loc[df["PD_OCUPA"] == "NAN", "PD_OCUPA"] = None

        # Normalizar ocupaciones
        df["PD_OCUPA"] = df["PD_OCUPA"].apply(normalize_ocupaciones)
    
    # --- AJUSTE DE ÍNDICES ---
    # Ajustar índices sumando 2 (para compensar encabezados en Excel)
    df.index = df.index + 2
    
    return df


# -------------------------
# Componentes de UI
# -------------------------
def setup_page_config():
    """Configura la página de Streamlit."""
    st.set_page_config(page_title="Monitoreo de Personas Desaparecidas", layout="wide")
    
    # Estilos personalizados
    st.markdown("""
    <style>
    div.stButton > button:first-child {
        background-color: #008CBA;
        color: black;
        border: none;
        padding: 10px 20px;
        text-align: center;
        text-decoration: none;
        display: inline-block;
        font-size: 16px;
        margin: 4px 2px;
        cursor: pointer;
        border-radius: 12px;
    }
    </style>
    """, unsafe_allow_html=True)

def create_date_filter(data):
    """Crea un filtro de fecha con botón de reinicio y opción para elegir entre
    filtrar por fecha de desaparición o fecha de denuncia."""
    
    # Verificar si existen las columnas requeridas
    has_fecdesap = "DD_FECDESAP" in data.columns
    has_fecden = "CI_FECDEN" in data.columns
    
    if not (has_fecdesap or has_fecden):
        return data.copy()
    
    # Inicializar variables para cada tipo de fecha
    fecha_min_desap = fecha_max_desap = None
    fecha_min_den = fecha_max_den = None
    
    # Calcular mínimos y máximos solo si existen las columnas (para evitar procesamiento innecesario)
    if has_fecdesap:
        fecha_min_desap = data["DD_FECDESAP"].min().date()
        fecha_max_desap = data["DD_FECDESAP"].max().date()
    
    if has_fecden:
        fecha_min_den = data["CI_FECDEN"].min().date()
        fecha_max_den = data["CI_FECDEN"].max().date()
    
    # Inicializar el estado si no existe
    if 'tipo_filtro_fecha' not in st.session_state:
        st.session_state['tipo_filtro_fecha'] = 'desaparicion'
    
    if 'rango_fechas_desap' not in st.session_state and has_fecdesap:
        st.session_state['rango_fechas_desap'] = [fecha_min_desap, fecha_max_desap]
    
    if 'rango_fechas_den' not in st.session_state and has_fecden:
        st.session_state['rango_fechas_den'] = [fecha_min_den, fecha_max_den]
    
    if 'date_input_key_desap' not in st.session_state:
        st.session_state['date_input_key_desap'] = 'filter_desap'
    
    if 'date_input_key_den' not in st.session_state:
        st.session_state['date_input_key_den'] = 'filter_den'
    
    # Selector de tipo de filtro
    tipo_filtro = st.sidebar.radio(
        "Filtrar por tipo de fecha:",
        options=['desaparicion', 'denuncia'],
        format_func=lambda x: "Fecha de Desaparición" if x == 'desaparicion' else "Fecha de Denuncia",
        key='radio_tipo_filtro'
    )
    
    # Actualizar el estado
    #st.session_state['tipo_filtro_fecha'] = tipo_filtro
    
    # Configuración del filtro según el tipo seleccionado
    col1, col2 = st.sidebar.columns([2, 1])
    
    filtered_data = data.copy()
    
    with col1:
        if tipo_filtro == 'desaparicion' and has_fecdesap:
            label = ":green[Rango] de :violet[Fechas de Desapariciones] para Métricas"
            rango_fechas = st.date_input(
                label,
                st.session_state['rango_fechas_desap'],
                key=st.session_state['date_input_key_desap']
            )
            
            if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
                start_date, end_date = rango_fechas
                # Usar .loc para mejor rendimiento
                mask = (data["DD_FECDESAP"].dt.date >= start_date) & (data["DD_FECDESAP"].dt.date <= end_date)
                filtered_data = data.loc[mask].copy()
                st.session_state['rango_fechas_desap'] = [start_date, end_date]
        
        elif tipo_filtro == 'denuncia' and has_fecden:
            label = ":green[Rango] de :blue[Fechas de Denuncias] para Métricas"
            rango_fechas = st.date_input(
                label,
                st.session_state['rango_fechas_den'],
                key=st.session_state['date_input_key_den']
            )
            
            if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
                start_date, end_date = rango_fechas
                # Usar .loc para mejor rendimiento
                mask = (data["CI_FECDEN"].dt.date >= start_date) & (data["CI_FECDEN"].dt.date <= end_date)
                filtered_data = data.loc[mask].copy()
                st.session_state['rango_fechas_den'] = [start_date, end_date]
    
    with col2:
        st.write("")
        if st.button("Reset", key="reset_date"):
            if tipo_filtro == 'desaparicion' and has_fecdesap:
                st.session_state['rango_fechas_desap'] = [fecha_min_desap, fecha_max_desap]
                st.session_state['date_input_key_desap'] = f'filter_desap_{datetime.now().strftime("%H%M%S")}'
            elif tipo_filtro == 'denuncia' and has_fecden:
                st.session_state['rango_fechas_den'] = [fecha_min_den, fecha_max_den]
                st.session_state['date_input_key_den'] = f'filter_den_{datetime.now().strftime("%H%M%S")}'
            
            # Incrementar la clave para forzar que se recreen los widgets
            if 'filter_key' in st.session_state:
                st.session_state.filter_key += 1
            st.rerun()
    
    return filtered_data

# -------------------------
# Funciones de análisis y visualización
# -------------------------
def display_metrics(df: pd.DataFrame):
    """Calcula y muestra las métricas clave."""
    fecha_min, fecha_max = df["DD_FECDESAP"].min().date(), df["DD_FECDESAP"].max().date()
    st.subheader(f"Metricas de {fecha_min} a {fecha_max}")

    total = len(df)
    localizados = df["PD_ESTATUSVIC"].value_counts().get("LOCALIZADA", 0)
    no_localizados = total - localizados
    promedio = df["PD_EDAD"].mean()

    col1, col2, col3, col4 = st.columns(4)
    col1.metric(":violet[Total Reportes]", total)
    col2.metric("Total :green[Localizados]", localizados, f"{localizados/total*100:.2f}%")
    col3.metric("Total :red[No Localizados]", no_localizados, f"{no_localizados/total*100:.2f}%")
    col4.metric(":orange[Edad Promedio]", f"{promedio:.1f} años")
    return total, localizados

def display_metrics_reaparecidos(df: pd.DataFrame):
    """Calcula y muestra las métricas clave relacionadas con las reapariciones, corrigiendo errores de digitación."""
    fecha_min, fecha_max = df["DD_FECDESAP"].min().date(), df["DD_FECDESAP"].max().date()
    st.subheader(f"Metricas de {fecha_min} a {fecha_max}")

    total = len(df)
    localizados = df["PD_ESTATUSVIC"].value_counts().get("LOCALIZADA", 0)

    # Calcular métricas de localizados si la columna existe
    localizados_con_vida = 0
    localizados_sin_vida = 0

    def normalizar_estatus(valor):
        if valor is None:
            return None
        
        # Convertir a minúsculas y eliminar espacios extra
        valor_norm = str(valor).lower().strip()
        
        # Mapeo de variaciones a valores estándar
        if 'con vida' in valor_norm: return 'CON VIDA'
        elif 'sin vida' in valor_norm: return 'SIN VIDA'
        return valor

    if "DL_LOCALSVCV" in df.columns and localizados > 0:
        df_loc = df[df["PD_ESTATUSVIC"] == "LOCALIZADA"]
        status_counts = df_loc["DL_LOCALSVCV"].value_counts().apply(normalizar_estatus)
        localizados_con_vida = status_counts.get("CON VIDA", 0)
        localizados_sin_vida = status_counts.get("SIN VIDA", 0)
        con_sin_vida = localizados_con_vida + localizados_sin_vida

    # Calcular el tiempo promedio de reaparición (en días), excluyendo outliers y errores de digitación
    if "DD_FECDESAP" in df.columns and "DL_FECLOC" in df.columns:
        df["TIEMPO_REAPARICION"] = (df["DL_FECLOC"] - df["DD_FECDESAP"]).dt.days

        # Filtrar outliers (1 año o más) y errores de digitación (tiempo negativo)
        df_filtrado = df[(df["TIEMPO_REAPARICION"] < 365) & (df["TIEMPO_REAPARICION"] >= 0)]

        tiempo_promedio_reaparicion = df_filtrado["TIEMPO_REAPARICION"].mean() if not df_filtrado.empty else np.nan
    else:
        st.error("Error con las columnas DD_FECDESAP y/o DL_FECLOC")
        tiempo_promedio_reaparicion = np.nan

    col1, col2, col3, col4 = st.columns(4)
    col1.metric(":blue[Total Localizados]", localizados, f"{localizados/total*100:.2f}%")
    col2.metric("Localizados :green[con Vida]", localizados_con_vida, f"{localizados_con_vida/localizados*100:.2f}%" if not np.isnan(localizados_con_vida) and localizados > 0 else "N/A")
    col3.metric("Localizados :red[sin Vida]", localizados_sin_vida, f"{localizados_sin_vida/localizados*100:.2f}%" if not np.isnan(localizados_sin_vida) and localizados > 0 else "N/A")
    col4.metric(":orange[Tiempo Promedio Localización (sin val. atípicos)]", f"{tiempo_promedio_reaparicion:.1f} días" if not np.isnan(tiempo_promedio_reaparicion) else "N/A", help="Sin valores atipicos: Sin casos mayores a 1 año de tiempo de localizacion")
    if localizados_con_vida + localizados_sin_vida != localizados:
        st.info(f"Hay :orange[{localizados - con_sin_vida}] casos en los que no se registró si se localizaron :green[CON VIDA] o :red[SIN VIDA]")    
    return total, localizados

@st.fragment
def plot_outliers_reapariciones(df: pd.DataFrame):
    """Visualiza los casos de reaparición de 1 año o más (outliers) y los errores de digitación."""
    st.markdown("---")
    st.subheader("Casos de :red[Reaparición Atípicos] y :red[Errores de Digitación]")
    if "DD_FECDESAP" in df.columns and "DL_FECLOC" in df.columns:
        # Convertir fechas y calcular tiempo de reaparición
        df["TIEMPO_REAPARICION"] = (df["DL_FECLOC"] - df["DD_FECDESAP"]).dt.days

        # Calcular tiempo en años
        df["TIEMPO_REAPARICION_ANOS"] = df["TIEMPO_REAPARICION"] / 365.25
        
        # Filtrar datos inválidos de una sola vez - valores de tiempo negativos o nulos
        df_validos = df.dropna(subset=["TIEMPO_REAPARICION", "DD_FECDESAP"])
        
        if df_validos.empty:
            st.warning("No hay datos de reaparición disponibles. Revise Etiquetas de Filtro")
            return
        
        # Extraer año y mes para operaciones posteriores
        df_validos["YEAR"] = df_validos["DD_FECDESAP"].dt.year
        df_validos["MONTH"] = df_validos["DD_FECDESAP"].dt.month
        df_validos["PERIODO_DESAPARICION"] = df_validos["DD_FECDESAP"].dt.to_period('M')
        
        # Obtener solo los años y meses que existen en los datos
        months_years = df_validos.groupby(["YEAR", "MONTH"]).size().reset_index()
        
        # Mapeo español para nombres de meses
        month_names = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        
        # Mapeo inverso para convertir nombres de meses de español a números
        month_numbers = {v: k for k, v in month_names.items()}
        
        # Crear opciones para el slider solo con los periodos existentes
        combined = []
        for _, row in months_years.sort_values(by=["YEAR", "MONTH"]).iterrows():
            combined.append(f"{month_names[row['MONTH']]} {row['YEAR']}")
        
        if not combined:
            st.warning("No hay periodos válidos disponibles.")
            return
        
        # Slider para seleccionar rango
        selected_range = st.select_slider("Rango de Meses y Años", combined, value=(combined[0], combined[-1]))
        
        # Procesar selección
        start_month_name, start_year = selected_range[0].split()
        end_month_name, end_year = selected_range[1].split()
        
        # Convertir nombres de meses a números
        start_month_num = month_numbers[start_month_name]
        end_month_num = month_numbers[end_month_name]
        
        # Convertir a fechas
        start_date = pd.Timestamp(year=int(start_year), month=start_month_num, day=1)
        if int(end_month_num) == 12:
            end_date = pd.Timestamp(year=int(end_year), month=end_month_num, day=31)
        else:
            end_date = pd.Timestamp(year=int(end_year), month=int(end_month_num) + 1, day=1) - pd.Timedelta(days=1)
        
        # Filtrar por rango de fechas seleccionado
        filtro = df_validos[(df_validos["DD_FECDESAP"] >= start_date) & (df_validos["DD_FECDESAP"] <= end_date)]

        # Outliers (1 año o más)
        df_outliers = filtro[filtro["TIEMPO_REAPARICION_ANOS"] >= 1]

        # Errores de digitación (tiempo negativo)
        df_errores = filtro[filtro["TIEMPO_REAPARICION"] < 0]

        if not df_outliers.empty:
            st.write("#### Casos de Reaparición de 1 Año o Más :red[(Atípicos)]")
            
            # Calcular el máximo valor para el eje x
            max_value = df_outliers["TIEMPO_REAPARICION_ANOS"].max()
            # Asegurar que el valor máximo sea al menos 40
            max_x = max(40, max_value)
            # Añadir 5% adicional para que se muestre el número 40
            max_x_display = max_x * 1.05
            
            # Crear el histograma con ajustes
            fig = px.histogram(
                df_outliers,
                x="TIEMPO_REAPARICION_ANOS",
                nbins=50,
                labels={
                    "TIEMPO_REAPARICION_ANOS": "Tiempo de Reaparición (años)",
                    "count": "Número de Casos"
                },
                title=f"Distribución del Tiempo de Reaparición (Outliers) - {len(df_outliers)} casos",
            )
            
            # Ajuste del eje x para que empiece en 1 y termine un poco después de 40
            fig.update_layout(
                xaxis=dict(
                    range=[1, max_x_display],  # Establece el rango mínimo en 1
                    dtick=5,                   # Establece las marcas principales cada 5 años
                ),
                yaxis=dict(
                    title="Número de Casos"     # Cambia la etiqueta del eje y
                )
            )
            
            # Ajuste de las barras para que estén correctamente alineadas
            fig.update_traces(
                xbins=dict(
                    start=1,          # Inicia las barras exactamente en 1
                    end=max_x_display,  # Establece el fin para incluir el valor 40
                    size=1            # Tamaño de cada barra (1 año)
                )
            )
            
            st.plotly_chart(fig, use_container_width=True)

            # Create two columns for the tables
            col1, col2 = st.columns(2)

            with col1:
                st.write("")
                st.write("")
                st.write(f"#### Detalles de :red[Valores Atípicos] - Total: {len(df_outliers)}")
                df_outliers_display = df_outliers[["DD_FECDESAP", "DL_FECLOC", "TIEMPO_REAPARICION"]]
                st.dataframe(df_outliers_display)

            with col2:
                if not df_errores.empty:
                    st.write(f"#### :red[Errores de Digitación] (Tiempo de Desaparición Negativo) - Total: {len(df_errores)}")
                    df_errores_display = df_errores[["DD_FECDESAP", "DL_FECLOC", "TIEMPO_REAPARICION"]]
                    st.dataframe(df_errores_display)
                else:
                    st.info("No se encontraron errores de digitación en las fechas en el rango seleccionado.")

        else:
            st.warning("No se encontraron casos de reaparición de 1 año o más en el rango seleccionado.")

            if not df_errores.empty:
                st.write(f"#### Errores de Digitación (Tiempo de Desaparición Negativo) - Total: {len(df_errores)}")
                df_errores_display = df_errores[["DD_FECDESAP", "DL_FECLOC", "TIEMPO_REAPARICION"]]
                st.dataframe(df_errores_display)
            else:
                st.info("No se encontraron errores de digitación en las fechas en el rango seleccionado.")
    else:
        st.warning("Las columnas 'DD_FECDESAP' y/o 'DL_FECLOC' no se encontraron.")

@st.fragment
def plot_desapariciones_por_año(df: pd.DataFrame):
    """
    Grafica la tendencia de desapariciones por año usando DD_FECDESAP.
    Muestra un slider para ajustar el intervalo visualizado.
    
    Args:
        df (pd.DataFrame): DataFrame con los datos de desapariciones
    
    Returns:
        None
    """
    if "DD_FECDESAP" not in df.columns or "PD_ESTATUSVIC" not in df.columns:
        st.warning("Columnas necesarias no encontradas.")
        return
    
    # Procesamiento de datos
    df["AÑO_DESAPARICION"] = df["DD_FECDESAP"].dt.year
    grupo = df.groupby(["AÑO_DESAPARICION", "PD_ESTATUSVIC"]).size().unstack(fill_value=0)
    
    # Aseguramos que ambas categorías estén presentes
    if "LOCALIZADA" not in grupo.columns:
        grupo["LOCALIZADA"] = 0
    if "EN INVESTIGACION" not in grupo.columns:
        grupo["EN INVESTIGACION"] = 0
    
    grupo["TOTAL"] = grupo.sum(axis=1)
    grupo.index = grupo.index.astype(int)
    grupo.index.name = "AÑO_DESAPARICION"
    
    # Aseguramos que todos los años estén representados
    all_years = range(grupo.index.min(), grupo.index.max() + 1)
    grupo = grupo.reindex(all_years, fill_value=0).reset_index()
    
    # Visualización
    st.subheader(" Registro (Histórico) de Desapariciones por Año")
    min_year, max_year = int(df["AÑO_DESAPARICION"].min()), int(df["AÑO_DESAPARICION"].max())
    selected_range = st.slider("Intervalo de años (Desapariciones)", min_year, max_year, (min_year, max_year))
    filtro = grupo[(grupo["AÑO_DESAPARICION"] >= selected_range[0]) & (grupo["AÑO_DESAPARICION"] <= selected_range[1])]

    fig = px.line(filtro, x="AÑO_DESAPARICION", y=["TOTAL", "LOCALIZADA", "EN INVESTIGACION"],
                  labels={"AÑO_DESAPARICION": "Año", "value": "Cantidad"},
                  markers=True,
                  color_discrete_map={"TOTAL": "blue", "LOCALIZADA": "green", "EN INVESTIGACION": "orange"})
    fig.update_layout(legend_title_text='')
    st.plotly_chart(fig, use_container_width=True)

####### ARIMA #######
def obtener_datos_censo():
    """Retorna los datos demográficos del censo para interpolación."""
    return {
        2010: {
            'total': 1955577, 'mujeres': 995389, 'hombres': 960188,
            'porcentajes': {
                'niños': 9.6, 'adolescentes_h': 5.0, 'adultos_h': 28.7, 'adultos_mayores_h': 5.8,
                'niñas': 9.3, 'adolescentes_m': 4.9, 'adultas_m': 30.1, 'adultas_mayores_m': 6.6
            }
        },
        2015: {
            'total': 2097175, 'mujeres': 1033907, 'hombres': 1063268,
            'porcentajes': {
                'niños': 9.2, 'adolescentes_h': 5.1, 'adultos_h': 29.0, 'adultos_mayores_h': 6.0,
                'niñas': 8.9, 'adolescentes_m': 5.0, 'adultas_m': 30.2, 'adultas_mayores_m': 6.6
            }
        },
        2020: {
            'total': 2320898, 'mujeres': 1148845, 'hombres': 1172053,
            'porcentajes': {
                'niños': 8.8, 'adolescentes_h': 5.2, 'adultos_h': 29.3, 'adultos_mayores_h': 6.2,
                'niñas': 8.5, 'adolescentes_m': 5.1, 'adultas_m': 30.3, 'adultas_mayores_m': 6.6
            }
        }
    }

def interpolar_valores(valor_antes, valor_despues, factor):
    """Función auxiliar para interpolación."""
    return int(valor_antes + factor * (valor_despues - valor_antes))

def generar_poblacion_por_año(año_inicio=2010, año_fin=None):
    """Genera población por año con desglose demográfico basado en datos de censo."""
    # Si no se especifica año_fin, usar hasta 6 años después del actual para predicciones
    if año_fin is None:
        año_fin = datetime.now().year + 6
        
    datos_censo = obtener_datos_censo()
    años_censo = sorted(datos_censo.keys())
    poblacion_por_año = []
    
    for año in range(año_inicio, año_fin + 1):
        if año in años_censo:
            # Usar datos exactos del censo
            censo = datos_censo[año]
            row = {
                'año': año,
                'poblacion_total': censo['total'],
                'mujeres': censo['mujeres'],
                'hombres': censo['hombres']
            }
            # Calcular población por categoría usando porcentajes
            for categoria, porcentaje in censo['porcentajes'].items():
                row[categoria] = int(censo['total'] * porcentaje / 100)
        else:
            # Determinar años de referencia para interpolación/extrapolación
            if año < años_censo[1]:
                año_antes, año_despues = años_censo[0], años_censo[1]
            elif año < años_censo[2]:
                año_antes, año_despues = años_censo[1], años_censo[2]
            else:
                # Extrapolación después de 2020
                tasa_crecimiento = ((datos_censo[2020]['total'] / datos_censo[2015]['total']) ** (1/5)) - 1
                años_extra = año - 2020
                censo_base = datos_censo[2020]
                
                poblacion_total = int(censo_base['total'] * ((1 + tasa_crecimiento) ** años_extra))
                row = {
                    'año': año,
                    'poblacion_total': poblacion_total,
                    'mujeres': int(censo_base['mujeres'] * ((1 + tasa_crecimiento) ** años_extra)),
                    'hombres': int(censo_base['hombres'] * ((1 + tasa_crecimiento) ** años_extra))
                }
                # Mantener proporciones de 2020
                for categoria, porcentaje in censo_base['porcentajes'].items():
                    row[categoria] = int(poblacion_total * porcentaje / 100)
                poblacion_por_año.append(row)
                continue
            
            # Interpolación
            factor = (año - año_antes) / (año_despues - año_antes)
            censo_antes, censo_despues = datos_censo[año_antes], datos_censo[año_despues]
            
            poblacion_total = interpolar_valores(censo_antes['total'], censo_despues['total'], factor)
            row = {
                'año': año,
                'poblacion_total': poblacion_total,
                'mujeres': interpolar_valores(censo_antes['mujeres'], censo_despues['mujeres'], factor),
                'hombres': interpolar_valores(censo_antes['hombres'], censo_despues['hombres'], factor)
            }
            
            # Interpolar porcentajes para categorías demográficas
            for categoria in censo_antes['porcentajes']:
                porcentaje_interpolado = (censo_antes['porcentajes'][categoria] + 
                                        factor * (censo_despues['porcentajes'][categoria] - censo_antes['porcentajes'][categoria]))
                row[categoria] = int(poblacion_total * porcentaje_interpolado / 100)
        
        poblacion_por_año.append(row)
    
    return pd.DataFrame(poblacion_por_año)

def clasificar_edad(edad):
    """Clasifica la edad en categorías demográficas."""
    if pd.isna(edad):
        return 'No especificado'
    elif edad < 12:
        return 'Niño'
    elif edad < 18:
        return 'Adolescente'
    elif edad < 60:
        return 'Adulto'
    else:
        return 'Adulto Mayor'

def calcular_poblacion_filtrada(df_poblacion, filtro_sexo, filtro_edad):
    """Calcula la población específica según los filtros aplicados."""
    # Mapeo de filtros a columnas
    mapeo_edad = {
        "Niño": ('niños', 'niñas'),
        "Adolescente": ('adolescentes_h', 'adolescentes_m'),
        "Adulto": ('adultos_h', 'adultas_m'),
        "Adulto Mayor": ('adultos_mayores_h', 'adultas_mayores_m')
    }
    
    if filtro_sexo == "Todos" and filtro_edad == "Todos":
        poblacion = df_poblacion['poblacion_total']
    elif filtro_sexo != "Todos" and filtro_edad == "Todos":
        poblacion = df_poblacion['hombres' if filtro_sexo == "Hombre" else 'mujeres']
    elif filtro_sexo == "Todos" and filtro_edad != "Todos":
        col_h, col_m = mapeo_edad[filtro_edad]
        poblacion = df_poblacion[col_h] + df_poblacion[col_m]
    else:
        col_h, col_m = mapeo_edad[filtro_edad]
        poblacion = df_poblacion[col_h if filtro_sexo == "Hombre" else col_m]
    
    return df_poblacion[['año']].assign(poblacion_total=poblacion)

def verificar_consistencia_poblacion():
    """Función para verificar que los porcentajes sumen correctamente."""
    datos_censo = obtener_datos_censo()
    
    print("Verificación de consistencia de porcentajes:")
    for año, datos in datos_censo.items():
        total_porcentaje = sum(datos['porcentajes'].values())
        total_calculado = datos['hombres'] + datos['mujeres']
        
        print(f"Año {año}: Suma de porcentajes = {total_porcentaje}%")
        print(f"  Total censo: {datos['total']:,}")
        print(f"  Hombres + Mujeres: {total_calculado:,}")
        print(f"  Diferencia: {abs(datos['total'] - total_calculado):,}")
        
        # Verificar suma por categorías de edad
        categorias_por_edad = {
            categoria: datos['total'] * sum_porcentajes / 100
            for categoria, sum_porcentajes in [
                ('Niños', datos['porcentajes']['niños'] + datos['porcentajes']['niñas']),
                ('Adolescentes', datos['porcentajes']['adolescentes_h'] + datos['porcentajes']['adolescentes_m']),
                ('Adultos', datos['porcentajes']['adultos_h'] + datos['porcentajes']['adultas_m']),
                ('Adultos Mayores', datos['porcentajes']['adultos_mayores_h'] + datos['porcentajes']['adultas_mayores_m'])
            ]
        }
        
        suma_categorias = sum(categorias_por_edad.values())
        print(f"  Suma por categorías de edad: {suma_categorias:,.0f}")
        print(f"  Diferencia con total: {abs(datos['total'] - suma_categorias):,.0f}\n")

def procesar_datos_para_arima(df):
    """Procesa los datos principales y genera las tasas por 100K habitantes con filtros demográficos."""
    
    # Selector de tipo de registro
    tipo_registro = st.radio("Seleccione sobre qué quiere hacer la predicción", 
                           ["Desapariciones", "Denuncias"], horizontal=True)
    
    # Selectores de filtros demográficos
    st.subheader("Filtros Demográficos")
    col1, col2, col3, col4 = st.columns([1,1,1,4])
    
    with col1:
        filtro_sexo = st.selectbox("Filtrar por Sexo:", 
                                  ["Todos", "Hombre", "Mujer"], 
                                  key="filtro_sexo")
    
    with col2:
        filtro_edad = st.selectbox("Filtrar por Clasificación de Edad:", 
                                  ["Todos", "Niño", "Adolescente", "Adulto", "Adulto Mayor"], 
                                  key="filtro_edad")
    
    with col3:
        filtro_estatus = st.selectbox("Filtrar por Estatus:", 
                                     ["Todos", "Localizada", "En Investigación"], 
                                     key="filtro_estatus")
    
    columna_fecha = "DD_FECDESAP" if tipo_registro == "Desapariciones" else "CI_FECDEN"
    
    # Validar columnas necesarias
    if columna_fecha not in df.columns:
        st.error(f"Columna necesaria no encontrada: {columna_fecha}")
        return pd.DataFrame(), 0, 0, pd.DataFrame()
    
    # Aplicar filtros demográficos al dataframe original
    filtros_aplicados = []
    
    # Filtro por sexo
    if filtro_sexo != "Todos":
        sexo_valor = "HOMBRE" if filtro_sexo == "Hombre" else "MUJER"
        df = df[df['PD_SEXO'] == sexo_valor]
        filtros_aplicados.append(f"Sexo: **{filtro_sexo}**")
    
    # Filtro por edad
    if filtro_edad != "Todos":
        df['CLASIFICACION_EDAD'] = df['PD_EDAD'].apply(clasificar_edad)
        df = df[df['CLASIFICACION_EDAD'] == filtro_edad]
        filtros_aplicados.append(f"Edad: **{filtro_edad}**")
    
    # Filtro por estatus
    if filtro_estatus != "Todos":
        estatus_valores = (["LOCALIZADA", "LOCALIZADA CON VIDA", "LOCALIZADA SIN VIDA"] 
                          if filtro_estatus == "Localizada" else ["EN INVESTIGACION"])
        df = df[df['PD_ESTATUSVIC'].isin(estatus_valores)]
        filtros_aplicados.append(f"Estatus: **{filtro_estatus}**")
    
    # Mostrar información de filtros aplicados
    with col4:
        if filtros_aplicados:
            col1_info, col2_info = st.columns([1,2])
            col2_info.info(f"Filtros aplicados: {', '.join(filtros_aplicados)}")
            col1_info.info(f"Registros después de aplicar filtros: **{len(df):,}**")
        else:
            st.write("")
            st.info("Sin filtros aplicados - usando todos los datos")

    # Generar dataset de población adaptativo
    año_actual = datetime.now().year
    año_fin_entrenamiento = año_actual - 1  # Hasta año anterior para entrenamiento
    año_fin_prediccion = año_actual + 8     # Hasta 8 años después para predicciones
    
    # Generar población para entrenamiento (hasta año anterior)
    df_poblacion_entrenamiento = generar_poblacion_por_año(año_fin=año_fin_entrenamiento)
    
    # Generar población completa para predicciones (hasta 6 años después del actual)
    df_poblacion_completo = generar_poblacion_por_año(año_fin=año_fin_prediccion)
    
    # Mostrar verificación en modo debug
    if st.checkbox("Mostrar verificación de consistencia de población"):
        verificar_consistencia_poblacion()
        st.dataframe(df_poblacion_entrenamiento)
    
    # Calcular población filtrada para entrenamiento
    df_poblacion = calcular_poblacion_filtrada(df_poblacion_entrenamiento, filtro_sexo, filtro_edad)
    
    # Procesar datos principales
    df["AÑO_REGISTRO"] = df[columna_fecha].dt.year
    
    # Filtrar datos desde 2010
    df = df[df["AÑO_REGISTRO"] >= 2010]
    
    if len(df) == 0:
        st.error("No hay datos disponibles después de aplicar los filtros seleccionados.")
        return pd.DataFrame(), 0, 0, pd.DataFrame()
    
    # Agrupar por año y combinar con población
    casos_por_año = df.groupby("AÑO_REGISTRO").size().reset_index(name='TOTAL')
    
    # Combinar con población filtrada para entrenamiento
    datos_completos = casos_por_año.merge(df_poblacion.rename(columns={'año': 'AÑO_REGISTRO'}), 
                                        on='AÑO_REGISTRO', how='left')
    
    # Llenar valores faltantes y calcular tasa
    datos_completos['poblacion_total'] = (datos_completos['poblacion_total']
                                        .interpolate()
                                        .fillna(method='bfill')
                                        .fillna(method='ffill')
                                        .replace(0, 1))
    
    datos_completos['TASA_POR_100K'] = (datos_completos['TOTAL'] / 
                                       datos_completos['poblacion_total'] * 100000)
    
    # Separar datos de entrenamiento (excluyendo año actual)
    training_data = (datos_completos[datos_completos['AÑO_REGISTRO'] < año_actual]
                    .sort_values('AÑO_REGISTRO')
                    .reset_index(drop=True))
    
    año_actual_data = datos_completos[datos_completos['AÑO_REGISTRO'] == año_actual]
    año_actual_valor = año_actual_data['TOTAL'].iloc[0] if not año_actual_data.empty else 0
    
    # Calcular población filtrada completa para predicciones
    df_poblacion_completo_filtrado = calcular_poblacion_filtrada(df_poblacion_completo, filtro_sexo, filtro_edad)
    
    poblacion_año_actual_row = df_poblacion_completo_filtrado[df_poblacion_completo_filtrado['año'] == año_actual]
    poblacion_año_actual = (poblacion_año_actual_row['poblacion_total'].iloc[0] 
                           if not poblacion_año_actual_row.empty else 0)
    
    st.success(f"Datos de entrenamiento: {len(training_data)} años (desde 2010 hasta {año_fin_entrenamiento})")
    
    return training_data, año_actual_valor, poblacion_año_actual, df_poblacion_completo_filtrado

@st.cache_resource
def entrenar_modelo_arima(training_data):
    """Entrena el modelo ARIMA con validaciones usando tasas normalizadas."""
    if len(training_data) < 3:
        st.warning("Se requieren al menos 3 años de datos para entrenar ARIMA.")
        return None
    
    # Usar tasas para el entrenamiento (mejor estabilidad estadística)
    serie_temporal = pd.to_numeric(training_data['TASA_POR_100K'], errors='coerce')
    serie_temporal = serie_temporal.replace([np.inf, -np.inf], np.nan).dropna()
    
    if len(serie_temporal) < 3 or serie_temporal.std() == 0:
        st.warning("Los datos no tienen suficiente variación para ARIMA.")
        return None
    
    try:
        with st.spinner("Entrenando modelo ARIMA..."):
            modelo = ARIMA(serie_temporal, order=(1, 1, 1)).fit()
        st.success("Modelo ARIMA entrenado exitosamente")
        return modelo
    except Exception as e:
        st.error(f"Error al entrenar ARIMA: {str(e)}")
        return None

@st.cache_data
def generar_predicciones(_modelo, training_data, df_poblacion, future_years=6, conf_level=0.95):
    """Genera predicciones con intervalos de confianza."""
    if _modelo is None:
        return pd.DataFrame()
    
    try:
        # Generar predicción en tasas
        forecast = _modelo.get_forecast(steps=future_years)
        conf_int = forecast.conf_int(alpha=1 - conf_level)
        forecast_values = forecast.predicted_mean
        
        # Años futuros
        ultimo_año = training_data['AÑO_REGISTRO'].max()
        años_futuros = range(ultimo_año + 1, ultimo_año + future_years + 1)
        
        # Obtener población futura
        poblaciones_futuras = []
        for año in años_futuros:
            pob_data = df_poblacion[df_poblacion['año'] == año]
            if not pob_data.empty:
                poblacion = pob_data['poblacion_total'].iloc[0]
            else:
                # Extrapolación simple si no hay datos
                ultima_poblacion = df_poblacion['poblacion_total'].iloc[-1]
                años_diff = año - df_poblacion['año'].max()
                crecimiento = 0.02  # 2% anual por defecto
                poblacion = ultima_poblacion * ((1 + crecimiento) ** años_diff)
            poblaciones_futuras.append(poblacion)
        
        # Convertir tasas a números absolutos para presentación
        predicciones_abs = forecast_values * np.array(poblaciones_futuras) / 100000
        limite_inf_abs = conf_int.iloc[:, 0] * np.array(poblaciones_futuras) / 100000
        limite_sup_abs = conf_int.iloc[:, 1] * np.array(poblaciones_futuras) / 100000
        
        # También convertir datos históricos a absolutos para continuidad visual
        datos_historicos_abs = training_data['TOTAL'].values
        
        # Crear DataFrame resultado
        df_pred = pd.DataFrame({
            'Año': años_futuros,
            'Poblacion_Proyectada': poblaciones_futuras,
            'Tasa_Predicha_por_100K': forecast_values,
            'Prediccion_Absoluta': np.maximum(0, predicciones_abs),
            'Limite_Inferior_Tasa': np.maximum(0, conf_int.iloc[:, 0]),
            'Limite_Superior_Tasa': conf_int.iloc[:, 1],
            'Limite_Inferior_Absoluto': np.maximum(0, limite_inf_abs),
            'Limite_Superior_Absoluto': limite_sup_abs
        })
        
        return df_pred
        
    except Exception as e:
        st.error(f"Error generando predicciones: {str(e)}")
        return pd.DataFrame()

def mostrar_graficas_mejoradas(predicciones, training_data, conf_level):
    """Muestra gráfica de casos absolutos."""
    st.subheader("📈 Visualización de Predicciones ARIMA")
    
    # GRÁFICA PRINCIPAL: Casos absolutos (más intuitiva para usuarios)
    fig = go.Figure()
    
    # Datos históricos absolutos
    fig.add_trace(go.Scatter(
        x=training_data['AÑO_REGISTRO'], 
        y=training_data['TOTAL'],
        mode='lines+markers', 
        name='Casos Históricos',
        line=dict(color='#2E86AB', width=3), 
        marker=dict(size=8),
        hovertemplate='<b>%{y:,.0f}</b> casos<br>Año: %{x}<extra></extra>'
    ))
    
    # Predicciones absolutas
    fig.add_trace(go.Scatter(
        x=predicciones['Año'], 
        y=predicciones['Prediccion_Absoluta'],
        mode='lines+markers', 
        name='Predicción de Casos',
        line=dict(color='#A23B72', width=3, dash='dash'), 
        marker=dict(size=8),
        hovertemplate='<b>%{y:,.0f}</b> casos<br>Año: %{x}<extra></extra>'
    ))
    
    # Intervalo de confianza absoluto
    fig.add_trace(go.Scatter(
        x=predicciones['Año'], 
        y=predicciones['Limite_Superior_Absoluto'],
        mode='lines', 
        line=dict(color='rgba(162,59,114,0.3)', width=1), 
        showlegend=False,
        hovertemplate='Límite superior: <b>%{y:,.0f}</b><extra></extra>'
    ))
    
    fig.add_trace(go.Scatter(
        x=predicciones['Año'], 
        y=predicciones['Limite_Inferior_Absoluto'],
        mode='lines', 
        name=f'Intervalo de Confianza ({conf_level*100:.0f}%)',
        line=dict(color='rgba(162,59,114,0.3)', width=1),
        fill='tonexty', 
        fillcolor='rgba(162,59,114,0.15)',
        hovertemplate='Límite inferior: <b>%{y:,.0f}</b><extra></extra>'
    ))
    
    fig.update_layout(
        title=f'<b>Predicción de Casos de Personas Desaparecidas</b><br><span style="font-size:12px">Intervalo de confianza: {conf_level*100:.0f}% | Entrenado con tasas normalizadas</span>',
        xaxis_title='Año', 
        yaxis_title='Número de Casos',
        hovermode='x unified', 
        template='plotly_white', 
        height=600,
        font=dict(size=12)
    )
    
    # Línea separadora
    ultimo_año_historico = training_data['AÑO_REGISTRO'].max()
    fig.add_vline(x=ultimo_año_historico + 0.5, line_dash="dot", line_color="gray",
                  annotation_text="Inicio de Predicción", annotation_position="top")
    
    st.plotly_chart(fig, use_container_width=True)
    

def mostrar_tabla_resultados_mejorada(predicciones, training_data, año_actual_valor, poblacion_año_actual):
    """Muestra tabla de resultados mejorada con énfasis en valores absolutos."""
    st.subheader("📋 Resultados de la Predicción")
    
    # Tabla principal con valores absolutos prominentes
    display_df = predicciones.copy()
    
    # Formatear para mejor visualización
    display_df['Población'] = display_df['Poblacion_Proyectada'].apply(lambda x: f"{x:,.0f}")
    display_df['Casos Esperados'] = display_df['Prediccion_Absoluta'].apply(lambda x: f"{max(0, x):.0f}")
    display_df['Rango Mínimo'] = display_df['Limite_Inferior_Absoluto'].apply(lambda x: f"{max(0, x):.0f}")
    display_df['Rango Máximo'] = display_df['Limite_Superior_Absoluto'].apply(lambda x: f"{x:.0f}")
    display_df['Tasa por 100K'] = display_df['Tasa_Predicha_por_100K'].apply(lambda x: f"{x:.2f}")
    
    col1, col2 = st.columns([2,1])
    with col1:
        # Tabla principal centrada en casos absolutos
        st.write("**🎯 Predicciones por Año:**")
        st.dataframe(
            display_df[['Año', 'Población', 'Casos Esperados', 'Rango Mínimo', 'Rango Máximo', 'Tasa por 100K']], 
            use_container_width=True,
            column_config={
                'Año': st.column_config.NumberColumn('📅 Año', format='%d'),
                'Población': st.column_config.TextColumn('👥 Población Proyectada'),
                'Casos Esperados': st.column_config.TextColumn('🎯 Casos Esperados'),
                'Rango Mínimo': st.column_config.TextColumn('📉 Mínimo Esperado'),
                'Rango Máximo': st.column_config.TextColumn('📈 Máximo Esperado'),
                'Tasa por 100K': st.column_config.TextColumn('📊 Tasa/100K')
            }
        )
    with col2:
        # Contexto actual
        año_actual = datetime.now().year
        ultimo_año_datos = training_data['AÑO_REGISTRO'].max()
        ultimo_año_casos = training_data.iloc[-1]['TOTAL']
        ultima_poblacion = training_data.iloc[-1]['poblacion_total']
        tasa_ultimo_año = (ultimo_año_casos / ultima_poblacion) * 100000 if ultima_poblacion > 0 else 0

        st.info(f"""
        **📅 Contexto de Datos:** 
        - **Último año con datos (Para entrenamiento):** {ultimo_año_datos} ({int(ultimo_año_casos):,} casos registrados)
        - **Población {ultimo_año_datos} (Proyección):** {ultima_poblacion:,.0f} habitantes
        - **Tasa {ultimo_año_datos}:** {tasa_ultimo_año:.2f} casos por cada 100K habitantes
        - **Año actual:** {año_actual} (primer año a predecir): {año_actual_valor} casos registrados
        - **Población {año_actual} (Proyección):** {poblacion_año_actual:,.0f} habitantes
        """)
    
    # Métricas clave mejoradas
    st.write("**📊 Métricas Clave:**")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        año_actual_pred = predicciones.iloc[0]['Prediccion_Absoluta']
        limite_inf = predicciones.iloc[0]['Limite_Inferior_Absoluto']
        limite_sup = predicciones.iloc[0]['Limite_Superior_Absoluto']
        año_actual = datetime.now().year
        st.metric(
            label=f"🎯 Año Actual ({año_actual})", 
            value=f"{año_actual_pred:.0f}",
            help=f"Rango esperado: {limite_inf:.0f} - {limite_sup:.0f} casos"
        )
    
    with col2:
        promedio_pred = predicciones['Prediccion_Absoluta'].mean()
        st.metric(
            label="📊 Promedio Anual", 
            value=f"{promedio_pred:.0f}",
            help="Promedio de casos esperados por año"
        )
    
    with col3:
        if len(training_data) > 0:
            ultimo_historico = training_data.iloc[-1]['TOTAL']
            año_actual_pred = predicciones.iloc[0]['Prediccion_Absoluta']
            cambio = ((año_actual_pred - ultimo_historico) / ultimo_historico * 100) if ultimo_historico > 0 else 0
            st.metric(
                label="📈 Cambio vs Último Año", 
                value=f"{cambio:+.1f}%",
                delta=f"{año_actual_pred - ultimo_historico:+.0f}"
            )
    
    with col4:
        total_periodo = predicciones['Prediccion_Absoluta'].sum()
        st.metric(
            label="🎯 Total Período", 
            value=f"{total_periodo:.0f}",
            help=f"Total esperado en {len(predicciones)} años"
        )

def mostrar_interpretacion_resultados(predicciones, training_data):
    """Muestra una sección de interpretación para ayudar al usuario."""
    with st.expander("🤔 ¿Cómo interpretar estos resultados?"):
        st.write("""
        ### 📖 Guía de Interpretación
        
        **🎯 Casos Esperados:**
        - Es el número más probable de casos que ocurrirán cada año
        - Se basa en patrones históricos normalizados por población
        - **El primer año mostrado es el año actual** (no hay datos reales aún)
        
        **📊 Intervalos de Confianza:**
        - **Rango Mínimo/Máximo:** Límites estadísticos esperados
        - Hay un 95% de probabilidad de que el valor real esté en este rango
        
        **🔬 Metodología:**
        - El modelo se entrena con **datos hasta el año anterior** al actual
        - Usa **tasas por 100K habitantes** para mayor precisión estadística
        - Los resultados se convierten a **números absolutos** para facilitar interpretación
        - Se considera el crecimiento poblacional proyectado
        
        **⚠️ Limitaciones:**
        - Las predicciones asumen que los patrones pasados continuarán
        - Eventos extraordinarios pueden alterar las predicciones  
        - La precisión disminuye a mayor distancia temporal
        - **La primera predicción es para el año en curso**, no para un año futuro
        """)

def mostrar_diagnostico(df, training_data, df_poblacion):
    """Muestra diagnóstico de datos mejorado."""
    st.subheader("🔍 Diagnóstico de Datos")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.write("**📊 Dataset Original:**")
        st.write(f"- Registros: {len(df):,}")
        st.write(f"- Años: {df['AÑO_REGISTRO'].min()} - {df['AÑO_REGISTRO'].max()}")
        st.write(f"- Rango temporal: {df['AÑO_REGISTRO'].max() - df['AÑO_REGISTRO'].min() + 1} años")
    
    with col2:
        st.write("**🎓 Datos de Entrenamiento:**")
        if not training_data.empty:
            st.write(f"- Años: {training_data['AÑO_REGISTRO'].min()} - {training_data['AÑO_REGISTRO'].max()}")
            st.write(f"- Registros: {len(training_data):,}")
            st.write(f"- Tasa promedio: {training_data['TASA_POR_100K'].mean():.2f}/100K")
            st.write(f"- Variabilidad: {training_data['TASA_POR_100K'].std():.2f}")
        else:
            st.write("- Sin datos disponibles")
    
    with col3:
        st.write("**👥 Dataset de Población:**")
        st.write(f"- Años: {df_poblacion['año'].min()} - {df_poblacion['año'].max()}")
        st.write(f"- Registros: {len(df_poblacion):,}")
        if len(df_poblacion) > 0:
            pob_min = df_poblacion['poblacion_total'].min()
            pob_max = df_poblacion['poblacion_total'].max()
            st.write(f"- Rango poblacional: {pob_min:,.0f} - {pob_max:,.0f}")
    
    with col4:
        
        st.write("**✅ Validaciones:**")
        # Información adicional sobre calidad de datos
        if not training_data.empty:
            if len(training_data) >= 3:
                st.success("✅ Datos suficientes para ARIMA")
            else:
                st.error("❌ Datos insuficientes (<3 años)")
        
            if training_data['TASA_POR_100K'].std() > 0:
                st.success("✅ Variabilidad adecuada")
            else:
                st.error("❌ Sin variabilidad en los datos")
        
            nulos = training_data['TASA_POR_100K'].isna().sum()
            if nulos == 0:
                st.success("✅ Sin valores faltantes")
            else:
                st.warning(f"⚠️ {nulos} valores faltantes")
    


# FUNCIÓN PRINCIPAL MEJORADA
@st.fragment
def ejecutar_prediccion_arima(data):
    """Función principal mejorada con mejor experiencia de usuario."""
    st.subheader("🔮 Predicción de Personas Desaparecidas y No Localizadas con ARIMA")
    st.write("**Predicciones basadas en análisis estadístico de patrones históricos normalizados por población**")
    
    # Controles de configuración
    col1, col2 = st.columns(2)
    with col1:
        confidence_level = st.slider("Nivel de Confianza", 0.80, 0.99, 0.95, 0.01, 
                                   help="Nivel de certeza estadística para los intervalos")
    with col2:
        years_to_predict = st.slider("Años a Predecir", 1, 10, 5, 1,
                                   help="Número de años futuros a proyectar")
    
    # Procesar datos
    resultado = procesar_datos_para_arima(data)
    if len(resultado) != 4:
        return
    
    training_data, año_actual_valor, poblacion_año_actual, df_poblacion = resultado
    
    # Diagnóstico opcional
    if st.checkbox("🔍 Mostrar diagnóstico de datos", value=False):
        mostrar_diagnostico(data, training_data, df_poblacion)
    
    # Validar datos suficientes
    if len(training_data) < 3 or training_data['TASA_POR_100K'].std() == 0:
        st.warning("⚠️ Datos insuficientes o sin variación para generar predicciones confiables.")
        if not training_data.empty:
            st.write("**Datos disponibles:**")
            st.dataframe(training_data[['AÑO_REGISTRO', 'TOTAL', 'TASA_POR_100K']])
        return
    
    # Entrenar y predecir
    modelo = entrenar_modelo_arima(training_data)
    if modelo is None:
        return
    
    predicciones = generar_predicciones(modelo, training_data, df_poblacion, years_to_predict, confidence_level)
    if predicciones.empty:
        return
    
    # Mostrar resultados mejorados
    mostrar_graficas_mejoradas(predicciones, training_data, confidence_level)
    mostrar_tabla_resultados_mejorada(predicciones, training_data, año_actual_valor, poblacion_año_actual)
    mostrar_interpretacion_resultados(predicciones, training_data)
    
    # Información técnica del modelo (expandible y optimizada)
    with st.expander("🔧 Información Técnica del Modelo"):
        st.markdown("### 🧠 Modelo ARIMA(1,1,1) para Desapariciones en Yucatán")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            **📊 Arquitectura del Modelo:**
            ```
            Xt = φ₁Xt-1 + θ₁εt-1 + εt
            ```
            - **AR(1):** Término autorregresivo de orden 1
            - **I(1):** Diferenciación de primer orden  
            - **MA(1):** Media móvil de orden 1
            - **Método:** Maximum Likelihood Estimation
            """)
            
            st.markdown("""
            **🎯 Metodología:**
            - **Input:** Tasas por 100K habitantes
            - **Output:** Casos absolutos proyectados
            - **Ventana:** Hasta año anterior para entrenamiento
            - **Conversión:** Tasa × Población proyectada / 100K
            """)
        
        with col2:
            st.markdown("""
            **📈 Análisis de la Serie Temporal:**
            """)
            
            if not training_data.empty:
                serie_tasas = training_data['TASA_POR_100K']
                
                # Métricas estadísticas compactas
                metrics_data = {
                    "Media": f"{serie_tasas.mean():.3f}",
                    "Desv. Estándar": f"{serie_tasas.std():.3f}",
                    "Coef. Variación": f"{(serie_tasas.std() / serie_tasas.mean()) * 100 if serie_tasas.mean() > 0 else 0:.1f}%",
                    "Tendencia": f"{training_data['AÑO_REGISTRO'].corr(serie_tasas):.3f}"
                }
                
                for metric, value in metrics_data.items():
                    st.write(f"• **{metric}:** {value}")
            
            st.markdown("""            
            **📊 Intervalos de Confianza:**
            - **Distribución:** Normal asintótica
            - **Cálculo:** Predicción ± Z × σ̂
            - **Característica:** Incrementan con horizonte temporal
            """)
        
        # Datos de entrenamiento optimizados
        st.markdown("### 📈 Datos Históricos Utilizados")
        
        # Formateo eficiente de datos
        display_cols = ['AÑO_REGISTRO', 'TOTAL', 'poblacion_total', 'TASA_POR_100K']
        display_training = training_data[display_cols].copy()
        
        # Aplicar formato solo a las columnas numéricas que lo necesitan
        format_funcs = {
            'poblacion_total': lambda x: f"{x:,.0f}",
            'TASA_POR_100K': lambda x: f"{x:.3f}",
            'TOTAL': lambda x: f"{x:,}"
        }
        
        for col, func in format_funcs.items():
            if col in display_training.columns:
                display_training[col] = display_training[col].apply(func)
        
        st.dataframe(
            display_training, 
            use_container_width=True,
            column_config={
                'AÑO_REGISTRO': st.column_config.NumberColumn('Año', format='%d'),
                'TOTAL': st.column_config.TextColumn('Casos'),
                'poblacion_total': st.column_config.TextColumn('Población'),
                'TASA_POR_100K': st.column_config.TextColumn('Tasa/100K')
            }
        )
        
        # Diagnóstico estadístico (condicional)
        if hasattr(modelo, 'summary') and st.checkbox("📊 Mostrar diagnóstico estadístico completo", value=False):
            with st.container():
                st.markdown("**Resumen del ajuste ARIMA:**")
                st.text(str(modelo.summary()))
        
        # Limitaciones técnicas específicas al contexto
        st.markdown("### ⚠️ Consideraciones Técnicas")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            **🚨 Limitaciones del Modelo:**
            - **Linealidad:** Asume patrones lineales temporales
            - **Horizonte:** Precisión decrece con el tiempo
            - **Variables exógenas:** No considera factores externos
            - **Estacionariedad:** Requiere patrones estables
            """)
        
        with col2:
            st.markdown("""
            **📊 Confiabilidad Temporal:**
            - **1-2 años:** Alta confiabilidad
            - **3-5 años:** Confiabilidad moderada  
            - **>5 años:** Usar con precaución
            - **Factores externos:** Pueden alterar predicciones
            """)
    st.markdown("---")
###############

@st.fragment
def plot_registros_por_año(df: pd.DataFrame):
    """
    Función unificada que grafica tanto el registro de denuncias como el de expedientes por año.
    Utiliza un selector de radio para alternar entre ambas visualizaciones,
    optimizando el uso de recursos para disminuir tiempos de carga.
    """
    st.subheader("📊 Análisis de Registros por Año")
    
    # Verificar columnas necesarias para ambas funcionalidades
    columnas_requeridas = ["CI_FECDEN", "PD_ESTATUSVIC", "CI_CARPEINV"]
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        st.warning(f"Columnas no encontradas: {', '.join(columnas_faltantes)}. Algunas funcionalidades podrían no estar disponibles.")
    
    # Selector de tipo de gráfica
    tipo_grafica = st.radio(
        "Seleccione el tipo de registro a visualizar:",
        ["Registro de Denuncias", "Registro de Expedientes"],
        horizontal=True
    )
    
    # Procesamiento inicial común para ambas gráficas
    # Como las columnas ya fueron limpiadas, asignamos directamente
    df_base = df
    
    if df_base.empty:
        st.warning("No hay datos de denuncias disponibles.")
        return
    
    # Añadir la columna de año a los datos base (será usada por ambas gráficas)
    df_base["AÑO_DENUNCIAS"] = df_base["CI_FECDEN"].dt.year
    
    # Filtro de años desde 2001 en adelante (común para ambas gráficas)
    df_base = df_base[df_base["AÑO_DENUNCIAS"] >= 2001]
    
    if df_base.empty:
        st.warning("No hay datos de denuncias desde 2001.")
        return
    
    # Determinar rango de años disponible para el slider
    min_year, max_year = int(df_base["AÑO_DENUNCIAS"].min()), int(df_base["AÑO_DENUNCIAS"].max())
    
    # Lógica específica según el tipo de gráfica seleccionado
    if tipo_grafica == "Registro de Denuncias":
        # Verificar columna específica para denuncias
        if "PD_ESTATUSVIC" not in df.columns:
            st.warning("Columna PD_ESTATUSVIC no encontrada. No se puede mostrar el registro de denuncias.")
            return
        
        # Crear pivot_table para denuncias
        conteo_denuncias = pd.pivot_table(
            df_base,
            index="AÑO_DENUNCIAS",
            columns="PD_ESTATUSVIC",
            aggfunc="size",
            fill_value=0
        ).reset_index()
        
        # Asegurar que existan las columnas necesarias
        cols = ["LOCALIZADA", "EN INVESTIGACION"]
        missing_cols = [col for col in cols if col not in conteo_denuncias.columns]
        if missing_cols:
            for col in missing_cols:
                conteo_denuncias[col] = 0
        
        # Calcular el total
        conteo_denuncias["TOTAL"] = conteo_denuncias[["LOCALIZADA", "EN INVESTIGACION"]].sum(axis=1)
        
        # UI para denuncias
        selected_range = st.slider(
            "Intervalo de años - Se toman los años en que se hicieron las denuncias",
            min_year, max_year, (min_year, max_year), 
            key="slider_denuncias"
        )
        
        # Filtrar según rango seleccionado
        mask_filtro = (conteo_denuncias["AÑO_DENUNCIAS"] >= selected_range[0]) & (conteo_denuncias["AÑO_DENUNCIAS"] <= selected_range[1])
        if mask_filtro.any():
            filtro = conteo_denuncias.loc[mask_filtro]
            fig = px.line(
                filtro, 
                x="AÑO_DENUNCIAS", 
                y=["TOTAL", "LOCALIZADA", "EN INVESTIGACION"],
                labels={"AÑO_DENUNCIAS": "Año", "value": "Cantidad"},
                markers=True,
                title="Tendencia de Denuncias por Año",
                color_discrete_map={"TOTAL": "purple", "LOCALIZADA": "green", "EN INVESTIGACION": "orange"}
            )
            fig.update_layout(legend_title_text='')
            st.plotly_chart(fig, use_container_width=True)
            
            # Añadir tabla de datos
            with st.expander("Ver datos detallados", expanded=False):
                st.dataframe(filtro, use_container_width=False)
                
                # Opción para descargar como CSV
                csv = filtro.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Descargar datos como CSV",
                    data=csv,
                    file_name="denuncias_por_año.csv",
                    mime="text/csv",
                )
        else:
            st.warning("No hay datos en el rango seleccionado.")
            
    else:  # Registro de Expedientes
        # Verificar columna específica para expedientes
        if "CI_CARPEINV" not in df.columns:
            st.warning("Columna CI_CARPEINV no encontrada. No se puede mostrar el registro de expedientes.")
            return
        
        # Como las columnas ya fueron limpiadas, usamos todos los datos
        df_exp = df_base
        
        if df_exp.empty:
            st.warning("No hay datos de expedientes disponibles.")
            return
        
        # Agrupar y contar expedientes por año
        conteo_expedientes = df_exp.groupby("AÑO_DENUNCIAS")["CI_CARPEINV"].nunique().reset_index(name="NUM_EXPEDIENTES")
        
        # UI para expedientes
        selected_range = st.slider(
            "Intervalo de años - Se toman los años en que se hicieron las denuncias",
            min_year, max_year, (min_year, max_year), 
            key="slider_expedientes"
        )
        
        # Filtrar según rango seleccionado
        mask_filtro = (conteo_expedientes["AÑO_DENUNCIAS"] >= selected_range[0]) & (conteo_expedientes["AÑO_DENUNCIAS"] <= selected_range[1])
        if mask_filtro.any():
            filtro = conteo_expedientes.loc[mask_filtro]
            fig = px.line(
                filtro, 
                x="AÑO_DENUNCIAS", 
                y="NUM_EXPEDIENTES",
                labels={"AÑO_DENUNCIAS": "Año", "NUM_EXPEDIENTES": "Número de Expedientes"},
                markers=True,
                title="Número de Expedientes de Desaparición por Año"
            )
            fig.update_layout(legend_title_text='')
            st.plotly_chart(fig, use_container_width=True)
            
            # Añadir tabla de datos
            with st.expander("Ver datos detallados", expanded=False):
                st.dataframe(filtro, use_container_width=False)

        else:
            st.warning("No hay datos en el rango de años seleccionado.")

def plot_sexo_distribution(df: pd.DataFrame):
    """Grafica la distribución por sexo (usando CI_FECDEN para filtrar años)."""
    st.subheader("👥 Distribución por Sexo")
    if "PD_SEXO" in df.columns and "CI_FECDEN" in df.columns:
        df["AÑO_DENUNCIAS"] = df["CI_FECDEN"].dt.year
        min_year = int(df["AÑO_DENUNCIAS"].min())
        max_year = int(df["AÑO_DENUNCIAS"].max())
        selected_range = st.slider("Intervalo de años (Sexo) - Se toman los años en que se hicieron las denuncias", min_year, max_year, (min_year, max_year), key="slider_sexo", help="Se toman los años de las denuncias")
        filtro = df[(df["AÑO_DENUNCIAS"] >= selected_range[0]) & (df["AÑO_DENUNCIAS"] <= selected_range[1])]
        counts = filtro["PD_SEXO"].value_counts()
        fig = px.pie(values=counts.values, names=counts.index,
                     title="Distribución de Denuncias por Sexo")
        fig.update_traces(textinfo="percent+label", insidetextorientation="radial")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("Columnas necesarias no encontradas.")

@st.fragment
def plot_clasedad_distribution(df: pd.DataFrame):
    """Grafica la distribución por clasificación etaria (usando CI_FECDEN)."""
    st.subheader("Distribución por Clasificación de Edad")
    
    if not {"PD_CLASEDAD", "CI_FECDEN"}.issubset(df.columns):
        st.warning("Columnas necesarias no encontradas.")
        return
    
    # Preparar datos
    df_work = df.copy()
    df_work["AÑO_DENUNCIAS"] = df_work["CI_FECDEN"].dt.year
    min_year, max_year = int(df_work["AÑO_DENUNCIAS"].min()), int(df_work["AÑO_DENUNCIAS"].max())
    
    # Filtro por años
    selected_range = st.slider("Intervalo de años (Clasificación de Edad) - Se toman los años de las denuncias", 
                              min_year, max_year, (min_year, max_year), 
                              help="Se toman los años de las denuncias", key="slider_clasedad")
    
    filtro = df_work[(df_work["AÑO_DENUNCIAS"] >= selected_range[0]) & 
                     (df_work["AÑO_DENUNCIAS"] <= selected_range[1])]
    
    # Gráfica
    counts = filtro["PD_CLASEDAD"].value_counts()
    pct = (counts / counts.sum()) * 100
    
    fig = px.bar(x=counts.index, y=counts.values,
                 labels={"x": "Clasificación de Edad", "y": "Cantidad"},
                 title="Distribución de Denuncias por Clasificación de Edad")
    
    for i, (label, count) in enumerate(counts.items()):
        fig.add_annotation(x=label, y=count, text=f"{count} ({pct.iloc[i]:.1f}%)", yanchor="bottom")
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Tabla
    st.subheader("Distribución de Clasificación de Edad por Año")
    st.warning('Verificar que no haya incongruencias entre "AÑO_CI" y "CI_FECDEN". Revisar tabla de errores al final de la Sección "No Localizados"')
    table_data = pd.crosstab(filtro["AÑO_DENUNCIAS"], filtro["PD_CLASEDAD"], margins=False, dropna=False)
    table_data["Total"] = table_data.sum(axis=1)
    table_data = table_data.reset_index()
    table_data = table_data.rename(columns={"AÑO_DENUNCIAS": "Año Denuncias"})
    
    # Reordenar columnas para que "NIÑO" esté después de "Año Denuncias"
    cols = list(table_data.columns)
    
    # --- Esto FALLABA ---- REVISAR --- #
    if "NIÑO" in cols:
        cols.remove("NIÑO")
        cols.insert(1, "NIÑO")
        table_data = table_data[cols]
    
    # Fila de totales
    totales = ["TOTALES"] + [table_data[col].sum() for col in table_data.columns[1:]]
    table_data.loc[len(table_data)] = totales
    table_data.index = range(1, len(table_data) + 1)
    
    st.dataframe(table_data, use_container_width=True)

@st.fragment
def plot_age_distribution(df: pd.DataFrame):
    """Grafica la distribución de edades con opciones para agrupar datos y muestra tabla de detalles."""
    st.subheader("📊 Distribución de Edades")
    
    # Verificar si existe la columna de edad
    if "PD_EDAD" not in df.columns:
        st.warning("Columna de edad no encontrada.")
        return
    
    # Preparar datos y aplicar filtros
    year_filter_applied = False

    # Aplicar filtros usando máscaras booleanas en lugar de crear copias
    if "CI_FECDEN" in df.columns:
        años_denuncias = df["CI_FECDEN"].dt.year
        min_year = int(años_denuncias.min())
        max_year = int(años_denuncias.max())
        selected_range = st.slider("Intervalo de años - Se toman los años de las denuncias", min_year, max_year, (min_year, max_year),
                                   help="Se toman los años de las denuncias", key="slider_edades")
        
        # Crear máscara para filtrar sin copiar el dataframe
        mask_años = (años_denuncias >= selected_range[0]) & (años_denuncias <= selected_range[1])
        year_filter_applied = True
    else:
        # Si no hay columna de fechas, usamos todos los datos
        mask_años = pd.Series(True, index=df.index)

    # Datos de edad para análisis
    # Extraer edades aplicando la máscara directamente
    edades = df.loc[mask_años, "PD_EDAD"].dropna()
    
    if len(edades) == 0:
        st.warning("No hay datos de edad disponibles para el filtro seleccionado.")
        return
        
    min_edad = int(edades.min())
    max_edad = int(edades.max())
    
    # Configuración de controles y estadísticas en columnas
    col1, col2 = st.columns([1, 1])
    
    # Columna 1: Controles de visualización
    with col1:
        # Opciones de visualización
        subcol1, subcol2 = st.columns([1, 1])
        with subcol1:
            intervalo_edad = st.number_input("Tamaño del intervalo", 
                                        min_value=1, max_value=10, value=5, 
                                        help="Define el rango de años para cada grupo")
        with subcol2:
            vista = st.radio("Tipo de visualización", 
                        ["Barras", "Área"], 
                        horizontal=True)
        
        # Crear los intervalos de edad una sola vez
        max_bin = max_edad + (intervalo_edad - max_edad % intervalo_edad) if max_edad % intervalo_edad != 0 else max_edad
        bins = list(range(min_edad, max_bin + intervalo_edad, intervalo_edad))
        labels = [f"{bins[i]}-{bins[i+1]-1}" for i in range(len(bins)-1)]
        
        # Calcular distribución de edades por intervalo
        edad_intervalo = pd.cut(edades, bins=bins, right=False, labels=labels)
        counts = edad_intervalo.value_counts().sort_index()
        total = len(edades)
        
        # Mostrar el rango más frecuente
        rango_max = counts.idxmax()
        count_max = counts.max()
        pct_max = (count_max/total*100)
        st.info(f"💡 El rango de edad más común es **{rango_max}** con **{count_max}** casos ({pct_max:.1f}% del total)")
        
        if year_filter_applied or len(edades) < len(df):
            st.warning("La ausencia de algunos registros de edad hace que el total de casos para esta grafica no coincida con el total general de reportes registrados.")
    
    # Columna 2: Estadísticas descriptivas
    with col2:
        st.write("**Estadísticas de edad**")
        
        # Primera fila de métricas
        metrics_row1 = st.columns(3)
        metrics_row1[0].metric("Edad promedio", f"{edades.mean():.1f} años")
        metrics_row1[1].metric("Mediana", f"{edades.median():.1f} años")
        metrics_row1[2].metric("Moda", f"{edades.mode().iloc[0]:.0f} años")
        
        # Segunda fila de métricas
        metrics_row2 = st.columns(3)
        metrics_row2[0].metric("Mínimo", f"{min_edad:.0f} años")
        metrics_row2[1].metric("Máximo", f"{max_edad:.0f} años")
        metrics_row2[2].metric("Total casos", f"{len(edades):,}")
    
    # Calcular porcentajes para anotaciones
    pct = (counts / total) * 100
    
    # Configuración base del gráfico
    fig_config = {
        "x": counts.index,
        "y": counts.values,
        "labels": {"x": "Rango de Edad", "y": "Cantidad"},
        "title": "Distribución de Edades",
        "color_discrete_sequence": ["#3366CC"]
    }
    
    # Crear el gráfico según el tipo seleccionado
    fig = px.bar(**fig_config) if vista == "Barras" else px.area(**fig_config)
    
    # Añadir línea de tendencia si es gráfico de barras
    if vista == "Barras":
        fig.add_scatter(x=counts.index, y=counts.values, mode='lines', 
                       line=dict(color='#FF6600', width=2), name='Tendencia')
    
    # Añadir anotaciones con valores y porcentajes
    for i, (idx, value) in enumerate(zip(counts.index, counts.values)):
        fig.add_annotation(
            x=idx, y=value,
            text=f"{value} ({pct.iloc[i]:.1f}%)",
            yanchor="bottom",
            showarrow=False,
            font=dict(size=10, color="black"),
            bgcolor="rgba(255, 255, 255, 0.7)"
        )
    
    # Estilo del gráfico
    fig.update_layout(
        xaxis_title="Rango de Edad",
        yaxis_title="Cantidad",
        hoverlabel=dict(bgcolor="blue"),
        margin=dict(l=20, r=20, t=50, b=20),
        plot_bgcolor="#8c8c8c",
        paper_bgcolor="#a9a9a9",
        font=dict(color="black", size=12)
    )
    
    # Configuración de ejes y cuadrícula
    for axis_update in [fig.update_xaxes, fig.update_yaxes]:
        axis_update(
            showgrid=True, 
            gridwidth=1, 
            gridcolor='#454546',
            title_font=dict(color="black", size=12),
            tickfont=dict(color="black", size=11)
        )
    
    # Mostrar el gráfico
    st.plotly_chart(fig, use_container_width=True)
    
    # Preparar y mostrar tabla de detalles
    st.markdown("#### Detalles de Registros")

    # Filtrar el dataframe original con la máscara de años
    df_filtrado = df.loc[mask_años]

    # Seleccionar las columnas requeridas
    columnas_tabla = ["index", "PD_EDAD", "PD_CLASEDAD", "PD_SEXO", "PD_ESTATUSVIC"]

    # Verificar si todas las columnas existen
    columnas_existentes = [col for col in columnas_tabla if col in df_filtrado.columns]
    
    if len(columnas_existentes) > 1:
        # Añadir el índice original como columna
        df_tabla = df_filtrado[columnas_existentes]
        
        # Renombrar columnas para mayor claridad
        nombre_columnas = {
            "index": "Índice Original",
            "PD_EDAD": "Edad",
            "PD_CLASEDAD": "Clasificación Edad",
            "PD_SEXO": "Sexo",
            "PD_ESTATUSVIC": "Estatus"
        }
        df_tabla.rename(columns=nombre_columnas, inplace=True)

        # Crear un expander para las tablas
        with st.expander("Detalles Completos"):
            # Dividir el expander en dos columnas
            col1, col2 = st.columns(2)
            
            # Columna 1: Tabla de Detalles
            with col1:
                personas_sin_edad = df_tabla["Edad"].isna().sum()
                st.write("### Tabla de Registros")
                st.write(f"Número de registros sin edad: {personas_sin_edad}")
                st.dataframe(df_tabla, use_container_width=True)
            
            # Columna 2: Tabla de Errores
            with col2:
                st.write("### :red[Errores] de Registro de Edad")
                # Validación de clasificación de edad
                df_errores = validar_clasificacion_edad(df_tabla)

                # Mostrar tabla de errores
                numero_errores = len(df_errores)    # Contar número de errores
                st.write(f"Total de registros con error: {numero_errores}")
                st.dataframe(df_errores, use_container_width=True)

    else:
        st.warning("No se encontraron todas las columnas requeridas para la tabla de detalles.")

###################################
# ---- ERRORES y VALIDACIONES ----#
###################################
def validar_clasificacion_edad(df):
    # Identificar errores de clasificación de edad directamente del DataFrame original
    df_errores = df[
        (df['Edad'] < 0) |  # Edad negativa
        ((df['Edad'] >= 0) & (df['Edad'] <= 11) & (df['Clasificación Edad'].str.lower() != 'niño')) |  # Niño
        ((df['Edad'] >= 12) & (df['Edad'] <= 17) & (df['Clasificación Edad'].str.lower() != 'adolescente')) |  # Adolescente
        ((df['Edad'] >= 18) & (df['Edad'] <= 59) & (df['Clasificación Edad'].str.lower() != 'adulto')) |  # Adulto
        ((df['Edad'] >= 60) & (df['Clasificación Edad'].str.lower() != 'adulto mayor'))  # Adulto Mayor
    ]
    
    return df_errores

@st.fragment
def validar_fechas_desaparicion(df):
    """
    Valida la cronología de fechas relacionadas con una desaparición.
    
    Parámetros:
    df (pandas.DataFrame): DataFrame con columnas de fechas a validar
    
    Retorna:
    tuple: (DataFrame de errores cronológicos, DataFrame de fechas incompletas, DataFrame de errores AÑO_CI)
    """
    
    # Seleccionar solo las columnas necesarias
    columnas_necesarias = [
        "PD_NOMBRECOMPLETOVICFORMULA", 'PD_ESTATUSVIC', 'PD_SEXO', 'PD_EDAD', 
        'DD_FECDESAP', 'DD_FECPERCA', 'CI_FECDEN', 'DL_FECLOC', 'AÑO_CI'
    ]
    
    df_vista = df[columnas_necesarias].copy()
    
    # Columnas de fechas para validación cronológica
    columnas_fecha = ['DD_FECDESAP', 'DD_FECPERCA', 'CI_FECDEN', 'DL_FECLOC']
    
    # Validar cronología usando vectorización
    def validar_cronologia_vectorizada(df_subset):
        fechas_array = df_subset[columnas_fecha].values
        mask_errores = []
        
        for row in fechas_array:
            # Filtrar fechas válidas (no NaT/NaN)
            fechas_validas = [f for f in row if pd.notna(f)]
            
            # Verificar secuencia cronológica
            if len(fechas_validas) >= 2:
                error = any(fechas_validas[i] > fechas_validas[i+1] for i in range(len(fechas_validas) - 1))
                mask_errores.append(error)
            else:
                mask_errores.append(False)
        
        return mask_errores
    
    # Aplicar validaciones usando máscaras booleanas
    mascara_errores = validar_cronologia_vectorizada(df_vista)
    
    # Validar AÑO_CI vs CI_FECDEN usando operaciones vectorizadas
    mascara_errores_ano_ci = (
        df_vista['AÑO_CI'].notna() & 
        df_vista['CI_FECDEN'].notna() & 
        (df_vista['AÑO_CI'] != df_vista['CI_FECDEN'].dt.year)
    )
    
    # Crear DataFrames de errores
    columnas_cronologicas = columnas_necesarias[:-1]  # Excluir AÑO_CI
    df_errores = df_vista[mascara_errores][columnas_cronologicas]
    
    columnas_ano_ci = ["PD_NOMBRECOMPLETOVICFORMULA", 'PD_ESTATUSVIC', 'PD_SEXO', 'PD_EDAD', 'CI_FECDEN', 'AÑO_CI']
    df_errores_año_ci = df_vista[mascara_errores_ano_ci][columnas_ano_ci]
    
    # Separar fechas incompletas usando operaciones vectorizadas
    mascara_fechas_incompletas = (
        (df_errores['DL_FECLOC'].dt.year == 1970) & 
        (df_errores['DL_FECLOC'].dt.month == 1) & 
        (df_errores['DL_FECLOC'].dt.day == 1)
    )
    
    df_fechas_incompletas = df_errores[mascara_fechas_incompletas]
    df_errores = df_errores[~mascara_fechas_incompletas]
    
    # Función unificada para formatear fechas
    def formatear_fechas_df(df):
        df_formateado = df.copy()
        
        # Formatear columnas de fecha
        for col in columnas_fecha:
            if col in df_formateado.columns:
                df_formateado[col] = df_formateado[col].dt.strftime('%Y-%m-%d')
        
        # Formatear AÑO_CI como entero
        if 'AÑO_CI' in df_formateado.columns:
            df_formateado['AÑO_CI'] = df_formateado['AÑO_CI'].astype('Int64')
        
        return df_formateado
    
    # Aplicar formateo
    df_errores_formateado = formatear_fechas_df(df_errores)
    df_fechas_incompletas_formateado = formatear_fechas_df(df_fechas_incompletas)
    df_errores_ano_ci_formateado = formatear_fechas_df(df_errores_año_ci)
    
    # Función optimizada para resaltar errores cronológicos
    def highlight_cronological_errors(row):
        styles = [''] * len(row)
        
        # Obtener fechas válidas con sus posiciones
        fechas_con_pos = []
        for col in columnas_fecha:
            if col in row.index:
                try:
                    fecha = pd.to_datetime(row[col])
                    if pd.notna(fecha):
                        fechas_con_pos.append((row.index.get_loc(col), fecha))
                except:
                    pass
        
        # Verificar errores cronológicos
        for i in range(len(fechas_con_pos) - 1):
            if fechas_con_pos[i][1] > fechas_con_pos[i+1][1]:
                styles[fechas_con_pos[i][0]] = 'color: red; font-weight: bold'
        
        return styles
    
    # Función para resaltar fechas incompletas
    def highlight_incomplete_dates(row):
        styles = [''] * len(row)
        if 'DL_FECLOC' in row.index and row['DL_FECLOC'] == '1970-01-01':
            styles[row.index.get_loc('DL_FECLOC')] = 'color: red; font-weight: bold'
        return styles
    
    # Función para crear archivo Excel con estilos
    def crear_excel_con_estilos(df_errores_formateado):
        buffer = io.BytesIO()
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores Cronológicos"
        
        # Escribir encabezados
        for col_idx, col_name in enumerate(df_errores_formateado.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Escribir datos y aplicar estilos
        for row_idx, (_, row) in enumerate(df_errores_formateado.iterrows(), 2):
            # Obtener fechas válidas con sus posiciones para esta fila
            fechas_con_pos = []
            for col in columnas_fecha:
                if col in row.index:
                    try:
                        fecha = pd.to_datetime(row[col])
                        if pd.notna(fecha):
                            col_idx = df_errores_formateado.columns.get_loc(col) + 1
                            fechas_con_pos.append((col_idx, fecha))
                    except:
                        pass
            
            # Determinar qué celdas deben estar en rojo
            celdas_rojas = set()
            for i in range(len(fechas_con_pos) - 1):
                if fechas_con_pos[i][1] > fechas_con_pos[i+1][1]:
                    celdas_rojas.add(fechas_con_pos[i][0])
            
            # Escribir valores y aplicar estilos
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Aplicar estilo rojo si es necesario
                if col_idx in celdas_rojas:
                    cell.font = Font(color="FF0000", bold=True)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    # Mostrar resultados con Streamlit
    resultados = [
        ("### :red[Errores] Cronológicos", df_errores_formateado, highlight_cronological_errors, "No se encontraron errores cronológicos."),
        ("### Solo Años - :red[Fechas incompletas]", df_fechas_incompletas_formateado, highlight_incomplete_dates, "No se encontraron fechas incompletas."),
        ("### :red[Errores] AÑO_CI vs CI_FECDEN", df_errores_ano_ci_formateado, None, "No se encontraron errores de congruencia entre AÑO_CI y CI_FECDEN .")
    ]
    
    for idx, (titulo, df_resultado, funcion_estilo, mensaje_vacio) in enumerate(resultados):
        st.write(f"{titulo}: {len(df_resultado)}")
        if not df_resultado.empty:
            if idx == 0:
                st.info("Es dificil identificar cual es la fecha erronea en un conjunto de fechas, porque, o ésa es la fecha erronea o las otras son erroneas.")
                st.warning("Si los registros en :red[rojo]  No son los erroneos, entonces los de la :red[derecha] serian los más probables que sean los :red[erroneos].")
            
            # Mostrar tabla con estilos
            if funcion_estilo:
                styled_df = df_resultado.style.apply(funcion_estilo, axis=1)
                st.dataframe(styled_df)
            else:
                st.dataframe(df_resultado)
            
            # Botón de descarga solo para la primera tabla (Errores Cronológicos)
            if idx == 0:
                try:
                    excel_data = crear_excel_con_estilos(df_errores_formateado)
                    st.download_button(
                        label="📥 Descargar Errores Cronológicos (Excel)",
                        data=excel_data,
                        file_name=f"errores_cronologicos_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Descarga la tabla de errores cronológicos con formato de celdas en rojo para errores"
                    )
                except Exception as e:
                    st.error(f"Error al crear el archivo Excel: {str(e)}")
        else:
            st.write(mensaje_vacio)
    
    return df_errores, df_fechas_incompletas, df_errores_año_ci

@st.fragment
def plot_municipio_estados_distribution(df: pd.DataFrame):
    """Grafica la distribución por municipio (usando CI_FECDEN)."""
    st.subheader('🏙️ Distribución por Municipio y Estado')

    # Verificar columnas necesarias
    columnas_requeridas = ['DD_ESTADO', 'DD_MPIO', 'AÑO_CI', 'DD_ESTADO', 'PD_ESTATUSVIC']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    if columnas_faltantes:
        st.warning('No se encontraron las columnas: ' + ', '.join(columnas_faltantes))
        return

    # Año de denuncias existente
    año_denuncias = 'AÑO_CI'
    min_year = int(df[año_denuncias].min())
    max_year = int(df[año_denuncias].max())
    
    # Obtener años seleccionados
    selected_range = st.slider("Intervalo de años (Municipio o Estado) - Se toman los años de las denuncias", 
                             min_year, max_year, (min_year, max_year), 
                             key="slider_municipio", 
                             help="Se toman los años de las denuncias")
    
    # Filtrar datos por año - usar query es más eficiente para filtros simples
    filtro = df.query(f"AÑO_DENUNCIAS >= {selected_range[0]} and AÑO_DENUNCIAS <= {selected_range[1]}")
    
    if filtro.empty:
        st.warning("No hay datos para el intervalo de años seleccionado.")
        return

    # Lista oficial de municipios de Yucatán (sin acentos) - definirla fuera para reutilizarla
    MUNICIPIOS_YUCATAN_UPPER = {
        'ABALA', 'ACANCEH', 'AKIL', 'BACA', 'BOKOBA', 'BUCTZOTZ', 'CACALCHEN',
        'CALOTMUL', 'CANSAHCAB', 'CANTAMAYEC', 'CELESTUN', 'CENOTILLO', 'CONKAL',
        'CUNCUNUL', 'CUZAMA', 'CHACSINKIN', 'CHANKOM', 'CHAPAB', 'CHEMAX',
        'CHICXULUB PUEBLO', 'CHICHIMILA', 'CHIKINDZONOT', 'CHOCHOLA', 'CHUMAYEL',
        'DZAN', 'DZEMUL', 'DZIDZANTUN', 'DZILAM DE BRAVO', 'DZILAM GONZALEZ',
        'DZITAS', 'DZONCAUICH', 'ESPITA', 'HALACHO', 'HOCABA', 'HOCTUN', 'HOMUN',
        'HUHI', 'HUNUCMA', 'IXIL', 'IZAMAL', 'KANASIN', 'KANTUNIL', 'KAUA',
        'KINCHIL', 'KOPOMA', 'MAMA', 'MANI', 'MAXCANU', 'MAYAPAN', 'MERIDA',
        'MOCOCHA', 'MOTUL', 'MUNA', 'MUXUPIP', 'OPICHEN', 'OXKUTZCAB', 'PANABA',
        'PETO', 'PROGRESO', 'QUINTANA ROO', 'RIO LAGARTOS', 'SACALUM', 'SAMAHIL',
        'SANAHCAT', 'SAN FELIPE', 'SANTA ELENA', 'SEYE', 'SINANCHE', 'SOTUTA',
        'SUCILA', 'SUDZAL', 'SUMA DE HIDALGO', 'TAHDZIU', 'TAHMEK', 'TEABO',
        'TECOH', 'TEKAL DE VENEGAS', 'TEKANTO', 'TEKAX', 'TEKIT', 'TEKOM',
        'TELCHAC PUEBLO', 'TELCHAC PUERTO', 'TEMAX', 'TEMOZON', 'TEPAKAN', 'TETIZ',
        'TEYA', 'TICUL', 'TIMUCUY', 'TINUM', 'TIXCACALCUPUL', 'TIXKOKOB',
        'TIXMEHUAC', 'TIXPEHUAL', 'TIZIMIN', 'TUNKAS', 'TZUCACAB', 'UAYMA',
        'UCU', 'UMAN', 'VALLADOLID', 'XOCCHEL', 'YAXCABA', 'YAXKUKUL', 'YOBAIN'
    }
    
    # Preparar filtro para convertirlo a mayúsculas una sola vez
    filtro["DD_MPIO_UPPER"] = filtro["DD_MPIO"].str.upper()
    
    # ---------- #
    opcion = st.radio("Selecciona el nivel de análisis:", ["Municipios", "Estados"])

    #filtro = df[(df["AÑO_DENUNCIAS"] >= selected_range[0]) & (df["AÑO_DENUNCIAS"] <= selected_range[1])]

    if opcion == "Municipios":  # Gráfica de los 10 municipios con más denuncias - usar Series.nlargest es eficiente
        top_municipios = filtro["DD_MPIO"].value_counts().nlargest(10)
        total_denuncias = len(filtro)

        if not top_municipios.empty:
            pct = (top_municipios / total_denuncias) * 100

            fig = px.bar(x=top_municipios.index, y=top_municipios.values,
                            labels={"x": "Municipio", "y": "Cantidad"},
                            title="Top 10 Municipios con más Denuncias")

            for i, (municipio, count) in enumerate(top_municipios.items()):
                fig.add_annotation(x=municipio, y=count,
                                    text=f"{count} ({pct.iloc[i]:.1f}%)",
                                    yanchor="bottom")

            st.plotly_chart(fig, use_container_width=True)

    elif opcion == "Estados":
        counts = filtro["DD_ESTADO"].value_counts().nlargest(10)
        total = filtro["DD_ESTADO"].value_counts().sum()
        pct = (filtro["DD_ESTADO"].value_counts() / total) * 100

        fig = px.bar(x=counts.index, y=counts.values,
                        labels={"x": "Estado", "y": "Cantidad"},
                        title="Top 10 Estados con más Denuncias")

        for i, count in enumerate(counts):
            fig.add_annotation(x=counts.index[i], y=count,
                                text=f"{count} ({pct.iloc[i]:.1f}%)", yanchor="bottom")

        st.plotly_chart(fig, use_container_width=True)

        # Guardar la tabla completa para su uso en el expander final
        counts_all = filtro["DD_ESTADO"].value_counts().reset_index()
        counts_all.columns = ["Estado", "Cantidad"]
        counts_all["Porcentaje"] = (counts_all["Cantidad"] / total) * 100
        counts_all["Porcentaje"] = counts_all["Porcentaje"].round(2)

        # Guardamos en session_state para usarla en el expander final
        st.session_state["tabla_estados"] = counts_all

    # ---------- #
    
    # Calcular agrupaciones de una sola vez para reusarlas
    # Usamos Pandas.crosstab que es más eficiente para contar combinaciones
    agg_data = pd.crosstab(
        index=[filtro["DD_ESTADO"], filtro["DD_MPIO"], filtro["DD_MPIO_UPPER"]], 
        columns=filtro["PD_ESTATUSVIC"],
        margins=False
    ).reset_index()
    
    # Renombrar y ajustar el DataFrame para asegurar que tenemos las columnas correctas
    if "LOCALIZADA" not in agg_data.columns:
        agg_data["LOCALIZADA"] = 0
    if "EN INVESTIGACION" not in agg_data.columns:
        agg_data["EN INVESTIGACION"] = 0
    
    # Calcular cantidades
    agg_data["Cantidad"] = agg_data["LOCALIZADA"] + agg_data["EN INVESTIGACION"]
    
    # Renombrar columnas
    agg_data.rename(columns={
        "DD_ESTADO": "Estado", 
        "DD_MPIO": "Municipio", 
        "DD_MPIO_UPPER": "Municipio_UPPER",
        "LOCALIZADA": "Localizadas",
        "EN INVESTIGACION": "En Investigacion"
    }, inplace=True)
    
    # Calcular porcentaje
    total = agg_data["Cantidad"].sum()
    agg_data["Porcentaje"] = (agg_data["Cantidad"] / total * 100).round(2)
    
    # Ordenar por cantidad descendente
    agg_data.sort_values("Cantidad", ascending=False, inplace=True)
    
    # Seleccionar columnas relevantes
    counts_all = agg_data[["Estado", "Municipio", "Municipio_UPPER", "Cantidad", "Localizadas", "En Investigacion", "Porcentaje"]]
    
    # Filtrar datos para Yucatán
    yucatan_data_all = counts_all[counts_all["Estado"] == "YUCATAN"].copy()
    
    # Separar municipios reales de localidades para Yucatán - usar vectorización
    yucatan_data_all["Es_Municipio_Real"] = yucatan_data_all["Municipio_UPPER"].isin(MUNICIPIOS_YUCATAN_UPPER)
    
    # Dividir los datos - usar filtros vectorizados es mucho más rápido que iterar
    yucatan_data = yucatan_data_all[yucatan_data_all["Es_Municipio_Real"]].drop(columns=["Municipio_UPPER", "Es_Municipio_Real"])
    yucatan_correcciones_df = yucatan_data_all[~yucatan_data_all["Es_Municipio_Real"]].drop(columns=["Municipio_UPPER", "Es_Municipio_Real"])
    
    # También limpiar la columna Municipio_UPPER de counts_all
    counts_all = counts_all.drop(columns=["Municipio_UPPER"])
    
    # Preparar listas de estados una sola vez
    estados = sorted(filtro["DD_ESTADO"].unique())
    estados_sin_yucatan = [estado for estado in estados if estado != "Yucatan"]
    
    # Contadores
    total_municipios = len(counts_all)
    total_estados = len(estados_sin_yucatan)
    
    # Inicializar estados de expanders
    for key in ["expander_municipios_abierto", "expander_estados_abierto", "expander_yucatan_abierto"]:
        if key not in st.session_state:
            st.session_state[key] = key == "expander_yucatan_abierto"
    
    # Expander para municipios corregidos de Yucatán - prioridad más alta
    with st.expander("Ver Municipios de Yucatán y correcciones", expanded=st.session_state.expander_yucatan_abierto):
        if not st.session_state.expander_yucatan_abierto:
            st.session_state.expander_yucatan_abierto = True
            
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader(f"Municipios de Yucatán ({len(yucatan_data)})")
            if not yucatan_data.empty:
                st.dataframe(yucatan_data, height=min(400, 100 + len(yucatan_data) * 35))
            else:
                st.info("No hay datos para municipios oficiales de Yucatán en el período seleccionado.")
                
        with col2:
            st.subheader(f"Correcciones para Municipios ({len(yucatan_correcciones_df)})")
            if not yucatan_correcciones_df.empty:
                st.dataframe(yucatan_correcciones_df, height=min(400, 100 + len(yucatan_correcciones_df) * 35))
            else:
                st.info("No hay localidades categorizadas incorrectamente como municipios.")
    
    # Expander para tablas principales
    with st.expander(f"Ver tabla completa de todos los Municipios - Puede incluir localidades ({total_municipios})", 
                     expanded=st.session_state.expander_municipios_abierto):
        if not st.session_state.expander_municipios_abierto:
            st.session_state.expander_municipios_abierto = True
        
        st.subheader("Todos los municipios")
        st.markdown("De Yucatan y de otros Estados")
        st.dataframe(counts_all)
            
    
    # Expander para estados - solo si hay estados aparte de Yucatán
    if estados_sin_yucatan:
        with st.expander(f"Ver Municipios por Estado ({total_estados} estados)", 
                        expanded=st.session_state.expander_estados_abierto):
            if not st.session_state.expander_estados_abierto:
                st.session_state.expander_estados_abierto = True
            
            # Mostrar tabla de resumen por estado
            if "tabla_estados" in st.session_state:
                tabla_estados = st.session_state["tabla_estados"]

                # Calcular el total de denuncias en otros estados (excluyendo Yucatán)
                total_otro_estados = tabla_estados[tabla_estados["Estado"] != "Yucatan"]["Cantidad"].sum()

                st.subheader(f"Resumen de denuncias por Estado (otros) - Total: {total_otro_estados:,} casos")
                st.dataframe(tabla_estados, use_container_width=False)
                st.markdown("---")

            num_columnas = 3
            
            # Usar caching para los datos de cada estado
            @st.cache_data
            def get_estado_data(estado, df):
                return df[df["Estado"] == estado].reset_index(drop=True)
            
            # Procesar cada estado en grupos de 3
            for i in range(0, len(estados_sin_yucatan), num_columnas):
                cols = st.columns(num_columnas)
                
                for j in range(num_columnas):
                    if i + j < len(estados_sin_yucatan):
                        estado = estados_sin_yucatan[i + j]
                        estado_data = get_estado_data(estado, counts_all)
                        
                        if not estado_data.empty:
                            with cols[j]:
                                st.subheader(f"{estado} ({len(estado_data)})")
                                st.dataframe(estado_data, height=min(350, 80 + len(estado_data) * 35))
                
                # Agregar separador solo entre filas, no al final
                if i + num_columnas < len(estados_sin_yucatan):
                    st.markdown("---")


@st.fragment
def plot_reapariciones(df: pd.DataFrame):
    """
    Muestra el tiempo de reaparición (días) de personas desaparecidas con filtro de meses y años.
    Opción para intervalos dinámicos (semanales, mensuales, trimestrales) y gráfica de barras, o histograma.
    Permite elegir entre fecha de desaparición o fecha de denuncia para calcular el tiempo.
    Incluye tabla dinámica con los datos de la gráfica actual.
    """
    st.subheader("Tiempo de Reaparición")
    
    # Verificar si existen las columnas necesarias
    required_columns = ["DD_FECDESAP", "CI_FECDEN", "DL_FECLOC"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        st.warning(f"Columnas no encontradas: {', '.join(missing_columns)}. Posible problema con Base de Datos")
        return
    
    # Selección de fecha para calcular el tiempo
    fecha_base = st.radio("Calcular tiempo desde:", ["Desaparicion", "Denuncia"], horizontal=True)
    fecha_columna = "DD_FECDESAP" if fecha_base == "Desaparicion" else "CI_FECDEN"
    
    # Proceso inicial de datos - solo una vez
    # Convertir fechas y calcular tiempo de reaparición
    df["TIEMPO_REAPARICION"] = (df["DL_FECLOC"] - df[fecha_columna]).dt.days
    
    # Filtrar datos inválidos de una sola vez - valores de tiempo negativos o nulos
    df_reap = df.dropna(subset=["TIEMPO_REAPARICION", fecha_columna])
    df_reap = df_reap[df_reap["TIEMPO_REAPARICION"] >= 0]
    
    if df_reap.empty:
        st.warning("No hay datos de reaparición disponibles. Revise Etiquetas de Filtro")
        return
    
    # Extraer año y mes una sola vez para operaciones posteriores
    df_reap["YEAR"] = df_reap[fecha_columna].dt.year
    df_reap["MONTH"] = df_reap[fecha_columna].dt.month
    df_reap["PERIODO_DESAPARICION"] = df_reap[fecha_columna].dt.to_period('M')
    
    # Obtener solo los años y meses que existen en los datos
    months_years = df_reap.groupby(["YEAR", "MONTH"]).size().reset_index()
    
    # Mapeo español para nombres de meses
    month_names = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    
    # Crear opciones para el slider solo con los periodos existentes
    combined = []
    for _, row in months_years.sort_values(by=["YEAR", "MONTH"]).iterrows():
        combined.append(f"{month_names[row['MONTH']]} {row['YEAR']}")
    
    if not combined:
        st.warning("No hay periodos válidos disponibles.")
        return
    
    # Slider para seleccionar rango
    selected_range = st.select_slider("Rango de Meses y Años", combined, value=(combined[0], combined[-1]),key="rango_plot_reapariciones")
    
    # Procesar selección
    start_month, start_year = selected_range[0].split()
    end_month, end_year = selected_range[1].split()
    
    # Obtener índices numéricos
    start_month_num = list(month_names.keys())[list(month_names.values()).index(start_month)]
    end_month_num = list(month_names.keys())[list(month_names.values()).index(end_month)]
    start_year = int(start_year)
    end_year = int(end_year)
    
    # Filtrar datos por el periodo seleccionado - usar los campos ya calculados
    filtro = df_reap[
        ((df_reap["YEAR"] > start_year) | ((df_reap["YEAR"] == start_year) & (df_reap["MONTH"] >= start_month_num))) &
        ((df_reap["YEAR"] < end_year) | ((df_reap["YEAR"] == end_year) & (df_reap["MONTH"] <= end_month_num)))
    ]
    
    # Checkbox para seleccionar el tipo de visualización
    dynamic_intervals = st.checkbox("Intervalos Dinámicos y Gráfica de Barras")
    
    if filtro.empty:
        st.warning("No hay datos en el rango seleccionado.")
        return
    
    # Título dinámico según la selección del usuario
    tiempo_descripcion = f"desde {fecha_base}"
    
    # Diccionario para almacenar los datos para la tabla dinámica
    table_data = {}
    
    if dynamic_intervals:
        # Configuración para gráfica de barras con intervalos dinámicos
        interval_type = st.selectbox("Tipo de Intervalo", ["Semanales de 7 días", "Mensuales de 30 días", "Trimestrales de 90 días"])
        
        interval_size = 7  # Default a semanal
        if interval_type == "Mensuales de 30 días":
            interval_size = 30
        elif interval_type == "Trimestrales de 90 días":
            interval_size = 90
        
        # Calcular máximo valor considerado (8 intervalos)
        max_value = 8 * interval_size
        
        # Crear bins y etiquetas para los intervalos
        bins = [i * interval_size for i in range(9)]
        labels = [f"{bins[i]}-{bins[i+1]-1} días" for i in range(8)]
        
        # Filtrar datos para el rango y crear la categoría
        datos_rango = filtro[filtro["TIEMPO_REAPARICION"] < max_value].copy()
        
        if datos_rango.empty:
            st.warning(f"No hay datos menores a {max_value} días en el rango seleccionado.")
            return
        
        # Asignar categorías de manera eficiente
        datos_rango["RANGO_TIEMPO"] = pd.cut(
            datos_rango["TIEMPO_REAPARICION"], 
            bins=bins, 
            right=False, 
            labels=labels
        )
        
        # Contar datos por categoría de manera eficiente
        counts = datos_rango["RANGO_TIEMPO"].value_counts().reindex(labels).fillna(0).astype(int)
        
        # Calcular estadísticas
        total = len(datos_rango)
        percentages = (counts / total * 100).round(1)
        
        # Preparar etiquetas
        label_text = [f"{count} ({percentage}%)" if count > 0 else "" 
                     for count, percentage in zip(counts.values, percentages)]
        
        # Crear gráfico
        fig = px.bar(
            x=counts.index, 
            y=counts.values,
            labels={'x': 'Rango de Tiempo', 'y': 'Cantidad'},
            title=f"Distribución del Tiempo de Reaparición {tiempo_descripcion} (8 Intervalos {interval_type})",
            text=label_text,
            color_discrete_sequence=['skyblue']
        )
        # Además, ajustar el margen superior para asegurar espacio para etiquetas
        fig.update_layout(margin=dict(t=80, b=50, l=50, r=50))
        fig.update_traces(textposition='outside')
        st.plotly_chart(fig, use_container_width=True)
        
        # Información sobre datos fuera del rango
        datos_fuera = filtro[filtro["TIEMPO_REAPARICION"] >= max_value]
        if not datos_fuera.empty:
            porcentaje_fuera = round((len(datos_fuera) / len(filtro) * 100), 1)
            st.info(f"Nota: {len(datos_fuera)} registros ({porcentaje_fuera}%) tienen un tiempo de reaparición mayor a {max_value} días y no se muestran en la gráfica.")
        
        # Preparar datos para la tabla dinámica
        table_data = {
            "Rango de Tiempo (días)": counts.index.tolist(),
            "Cantidad": counts.values.tolist(),
            "Porcentaje (%)": percentages.tolist()
        }
        
    else:
        # Histograma original para tiempos menores a 30 días
        max_days_for_histogram = 29
        datos_histograma = filtro[filtro["TIEMPO_REAPARICION"] <= max_days_for_histogram]
        
        if datos_histograma.empty:
            st.warning(f"No hay datos menores a {max_days_for_histogram+1} días en el rango seleccionado.")
            return
        
        # Calcular el histograma de manera eficiente
        counts, bins = np.histogram(datos_histograma["TIEMPO_REAPARICION"], bins=30, range=(0, max_days_for_histogram+1))
        
        # Crear la figura con datos pre-calculados
        fig = px.histogram(
            datos_histograma, 
            x="TIEMPO_REAPARICION", 
            nbins=30,
            range_x=[0, max_days_for_histogram+1],
            labels={"TIEMPO_REAPARICION": "Días"},
            title=f"Distribución del Tiempo de Reaparición {tiempo_descripcion} (hasta 30 días)"
        )
        # Actualizar explícitamente el nombre del eje y
        fig.update_layout(yaxis_title="Cantidad",
                          margin = dict(t=80,b=50,l=50,r=50))
        
        # Agregar anotaciones eficientemente
        total_data_points = len(datos_histograma)
        for i in range(len(counts)):
            bin_center = (bins[i] + bins[i+1]) / 2
            if counts[i] > 0:  # Solo agregar anotaciones para barras con datos
                percentage = (counts[i] / total_data_points) * 100
                fig.add_annotation(
                    x=bin_center, 
                    y=counts[i],
                    text=f"{counts[i]}", 
                    yanchor="top", 
                    yshift=0
                )
                fig.add_annotation(
                    x=bin_center, 
                    y=counts[i],
                    text=f"{percentage:.1f}%", 
                    yanchor="bottom", 
                    yshift=25, 
                    showarrow=False
                )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Información sobre datos excluidos
        datos_excluidos = filtro[filtro["TIEMPO_REAPARICION"] > max_days_for_histogram]
        if not datos_excluidos.empty:
            porcentaje_excluidos = round((len(datos_excluidos) / len(filtro) * 100), 1)
            st.info(f"Nota: {len(datos_excluidos)} registros ({porcentaje_excluidos}%) tienen un tiempo de reaparición mayor a {max_days_for_histogram} días y no se muestran en el histograma.")
        
        # Preparar datos para la tabla dinámica - para el caso de histograma
        # Crear rangos más comprensibles para la tabla
        day_ranges = [f"{int(bins[i])}-{int(bins[i+1])-1}" for i in range(len(bins)-1)]
        percentages = [(count / total_data_points * 100).round(1) for count in counts]
        
        table_data = {
            "Rango de Tiempo (días)": day_ranges,
            "Cantidad": counts.tolist(),
            "Porcentaje (%)": percentages
        }
    
    # Crear expander para la tabla dinámica
    with st.expander("Ver tabla de datos", expanded=False):
        # Convertir los datos en un DataFrame para la tabla
        if table_data:
            df_table = pd.DataFrame(table_data)
            
            # Añadir una columna para el total acumulado y porcentaje acumulado
            df_table["Cantidad Acumulada"] = df_table["Cantidad"].cumsum()
            df_table["Porcentaje Acumulado (%)"] = (df_table["Cantidad Acumulada"] / df_table["Cantidad"].sum() * 100).round(1)
            
            # Formatear columnas numéricas como enteros donde corresponda
            df_table["Cantidad"] = df_table["Cantidad"].astype(int)
            df_table["Cantidad Acumulada"] = df_table["Cantidad Acumulada"].astype(int)
            
            # Mostrar la tabla
            st.write(f"Datos detallados del Tiempo de Reaparición {tiempo_descripcion}")
            st.dataframe(df_table, use_container_width=False)
            

# ---------------------- ANALISIS DE CONTEXTO---------------------- #
# ---------------------------
# Configuración global
# ---------------------------
CATEGORIAS_PREDEFINIDAS = {
    'CON_PAREJA': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'fuga': ['FUG', 'ESCAP', 'FUE CON', 'SE FUE', 'HUY', 'ESTA', 'IBA CON', 'IRIA CON'],
            'pareja': ['NOVIO', 'PAREJA', 'NOVIA', 'AMANTE', 'ENAMORAD', 'QUERID']
        }
    },
    'CON_AMIGOS': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'fuga': ['FUG', 'ESCAP', 'FUE CON', 'SE FUE', 'HUY', 'ESTA', 'IBA CON', 'IRIA CON'],
            'amistad': ['AMIG', 'AMISTAD']
        }
    },
    'CONYUGE': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'pareja': ['ESPOS', 'CONYUGE', 'MARIDO']
        }
    },
    'FUGA_DE_ALBERGUE': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'fuga': ['FUG', 'ESCAP', 'SALIO', 'NO REGRES', 'SE FUE', 'HUY'],
            'institucion': ['ALBERGUE', 'HORFANATO', 'CAIMEDE', 'REFUGIO', 'HOGAR', 'CONVENTO', 'DIF ', 'CRIA ']
        }
    },
    'INGRESO_ALBERGUE': {
        'grupo': 'TIPO_DESAPARICION_POSIBLE_CAUSA',
        'palabras_clave': {
            'fuga': ['ENTR', 'INGRES', 'INTERN'],
            'institucion': ['ALBERGUE', 'HORFANATO', 'CAIMEDE', 'REFUGIO', 'HOGAR', 'CONVENTO', 'DIF ', 'CRIA ']
        }
    },
    'SALIO_DE_CASA_NO_VOLVIO': {
        'grupo': 'HECHOS',
        'palabras_clave': {
            'fuga': ['FUG', 'ESCAP', 'SALIO', 'NO REGRES', 'SE FUE', 'HUY'],
            'casa': ['CASA', 'DOMICILIO', 'FAMILIA', 'ADOLESCENTE']
        }
    },
    'ADOLESCENTE_ESCAPA': {
        'grupo': 'HECHOS',
        'palabras_clave': {
            'casa': ['ADOLESCENTE QUE SE ESCAPA']
        }
    },
    'DISCUSION': {
        'grupo': 'HECHOS_PREVIOS_POSIBLE_CAUSA',
        'palabras_clave': {
            'problema': ['DISCU', 'PELE', 'GRIT', 'ARGUMENTO', 'PROBLEMA', 'MALTRAT', 'VIOLENCIA']
        }
    },
    'CRISIS_MENTAL': {
        'grupo': 'HECHOS_PREVIOS_POSIBLE_CAUSA',
        'palabras_clave': {
            'crisis': ['CRISIS', 'DEPRESION', 'NERVIOS', 'ANSIEDAD', 'TRASTORNO', 'ALZHEIMER', 'DEMENCIA', 'MENTAL', 'PADEC']
        }
    },
    'CENTRO_REHABILITACION': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'centro': ['REHABILITACION', 'ANEXO']
        }
    },
    'ESTABA_PADRE_MADRE': {
        'grupo': 'HECHOS',
        'palabras_clave': {
            'llevar': ['ESTA CON SU', 'ESTABA CON'],
            'familia': ['SU PAPA', 'SU PADRE', 'SU MAMA', 'SU MADRE']
        }
    },
    'LLEVADO_POR_PADRE_MADRE': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'llevar': ['SE LA LLEV', 'SE LO LLEV', 'SE LOS LLEV', 'SE LAS LLEV', 'CON SU HIJ', 'CON SUS HIJ', 'NIÑOS'],
            'familia': ['SU PAPA', 'SU PADRE', 'PAREJA', 'SU MAMA', 'SU MADRE', 'ELLA']
        }
    },
    'ABANDONO_HOGAR': {
        'grupo': 'TIPO_DESAPARICION',
        'palabras_clave': {
            'casa': ['ABANDONO DE HOGAR']
        }
    },
    'OTRO_ESTADO': {
        'grupo': 'HECHOS',
        'palabras_clave': {
            'estado': ["AGUASCALIENTES", "BAJA CALIFORNIA", "BAJA CALIFORNIA SUR", "CAMPECHE", "COAHUILA", "COLIMA", "CHIAPAS",
                       "CHIHUAHUA", "CIUDAD DE MEXICO", "DURANGO", "GUANAJUATO", "GUERRERO", "HIDALGO", "JALISCO", "ESTADO DE MEXICO",
                       "MICHOACAN", "MORELOS", "NAYARIT", "NUEVO LEON", "OAXACA", "PUEBLA", "QUERETARO", "QUINTANA ROO", "SAN LUIS POTOSI",
                       "SINALOA", "SONORA", "TABASCO", "TAMAULIPAS", "TLAXCALA", "VERACRUZ", "ZACATECAS"]
        }
    },
    'ACTIVIDAD': {
        'grupo': 'HECHOS_PREVIOS',
        'palabras_clave': {
            'actividad': ['SALIO A', 'FUE A', 'IRIA A', 'IBA A', 'RUMBO A', 'VA A', 'TRABAJ', 'ESTUDIAR', 'ESCUELA', 'UNIVERSIDAD', 'ACTIVIDAD', 'MONTAR', 'CORRER']
        }
    },
    'DETENCION': {
        'grupo': 'TIPO_DESAPARICION_POSIBLE_CAUSA',
        'palabras_clave': {
            'actividad': ['ESTAB', 'FUE', 'ESTUVO', 'IBA', 'LLEVAN', 'ENCONTRABA', 'DELITO'],
            'detencion': ['DETENID', 'ARRESTAD', 'RECLUID', 'CERESO']
        }
    },
    'EBRIEDAD_DROGAS': {
        'grupo': 'HECHOS_PREVIOS_POSIBLE_CAUSA',
        'palabras_clave': {
            'sustancias': ['ALCOHOL', 'EBRI', 'BORRACH', 'EMBRIAG', 'BEBID', 'TOMANDO', 'ALCO',
                            'DROGA', 'DROGAD', 'SUSTANCIA', 'ESTUPEFACIENTE', 'INTOXICA', 'BEBIENDO',
                            'MARIHUANA', 'COCAINA', 'PASTILLAS', 'CRACK', 'METANFETAMINA']
        }
    }
}

GRUPOS_CATEGORIAS = {info['grupo'] for info in CATEGORIAS_PREDEFINIDAS.values()}
COLORES_GRUPOS = {
    'TIPO_DESAPARICION': '#4E79A7',
    'TIPO_DESAPARICION_POSIBLE_CAUSA': '#F28E2B',
    'HECHOS': '#E15759',
    'HECHOS_PREVIOS': '#76B7B2',
    'HECHOS_PREVIOS_POSIBLE_CAUSA': '#59A14F'
}

# ---------------------------
# Funciones de Procesamiento
# ---------------------------
def categorizar_registros(df: pd.DataFrame) -> pd.DataFrame:
    """
    Añade columnas de categorías y grupos a partir de DD_HECHOS.
    """
    if df is None or df.empty:
        return None

    df_categorizado = df.copy()
    df_categorizado['DD_HECHOS'] = df_categorizado['DD_HECHOS'].astype(str).str.upper()

    # Inicializar columnas de categorías y grupos
    for categoria in CATEGORIAS_PREDEFINIDAS.keys():
        df_categorizado[categoria] = 0
    for grupo in GRUPOS_CATEGORIAS:
        df_categorizado[f'GRUPO_{grupo}'] = 0

    # Procesar cada registro
    for idx, row in df_categorizado.iterrows():
        descripcion = row['DD_HECHOS']
        for categoria, info in CATEGORIAS_PREDEFINIDAS.items():
            palabras_clave = info['palabras_clave']
            if all(any(re.search(r'\b' + palabra, descripcion) for palabra in lista)
                   for lista in palabras_clave.values()):
                df_categorizado.at[idx, categoria] = 1
                df_categorizado.at[idx, f'GRUPO_{info["grupo"]}'] = 1

    df_categorizado['TOTAL_CATEGORIAS'] = df_categorizado[list(CATEGORIAS_PREDEFINIDAS.keys())].sum(axis=1)
    df_categorizado['CATEGORIAS_DETECTADAS'] = df_categorizado.apply(
        lambda row: ', '.join([cat for cat in CATEGORIAS_PREDEFINIDAS.keys() if row[cat] == 1]) or 'SIN_CATEGORIA',
        axis=1
    )
    return df_categorizado

# ---------------------------
# Funciones de Visualización Interactiva con Plotly
# ---------------------------
def plot_resumen_categorias(df_categorizado: pd.DataFrame):
    resumen = df_categorizado[list(CATEGORIAS_PREDEFINIDAS.keys())].sum().sort_values(ascending=False).reset_index()
    resumen.columns = ['Categoria', 'Total']
    resumen['Grupo'] = resumen['Categoria'].apply(lambda cat: CATEGORIAS_PREDEFINIDAS[cat]['grupo'])
    fig = px.bar(resumen, x='Categoria', y='Total', color='Grupo',
                 text='Total', color_discrete_map=COLORES_GRUPOS)
    fig.update_layout(title="Distribución de registros por categoría", xaxis_tickangle=-45)
    return fig


def plot_comparativa_grupos(df_categorizado: pd.DataFrame):
    conteo = {grupo: df_categorizado[f'GRUPO_{grupo}'].sum() for grupo in GRUPOS_CATEGORIAS}
    data = pd.DataFrame({
        'Grupo': list(conteo.keys()),
        'Total': list(conteo.values())
    })
    fig = px.bar(data, x='Grupo', y='Total', text='Total', color='Grupo',
                 color_discrete_map=COLORES_GRUPOS)
    fig.update_layout(title="Comparativa de registros por grupo de categorías", xaxis_tickangle=-45)
    return fig

def display_metrics_contexto(df: pd.DataFrame):
    """Muestra métricas clave sobre la categorización de registros, incluyendo totales y porcentajes, con fechas de intervalo."""

    if "CI_FECDEN" not in df.columns or "TOTAL_CATEGORIAS" not in df.columns:
        st.error("El DataFrame debe contener las columnas 'CI_FECDEN' y 'TOTAL_CATEGORIAS'.")
        return

    # Calcular intervalo de fechas
    fecha_min, fecha_max = df["CI_FECDEN"].min().date(), df["CI_FECDEN"].max().date()
    st.subheader(f"Métricas de {fecha_min} a {fecha_max}")

    total = len(df)
    con_cat = (df["TOTAL_CATEGORIAS"] > 0).sum()
    sin_cat = (df["TOTAL_CATEGORIAS"] == 0).sum()

    porcentaje_con = (con_cat / total) * 100 if total > 0 else 0
    porcentaje_sin = (sin_cat / total) * 100 if total > 0 else 0

    col1, col2, col3 = st.columns(3)
    col1.metric(":blue[Total Registros]", total)
    col2.metric(":green[Con Categoría]", con_cat, f"{porcentaje_con:.2f}%")
    col3.metric(":red[Sin Categoría]", sin_cat, f"{porcentaje_sin:.2f}%")

    return total, con_cat, sin_cat



def plot_heatmap_categorias_por_grupo(df_categorizado: pd.DataFrame):
    # Crear matriz: filas=grupo, columnas=categoría
    data = {}
    for grupo in GRUPOS_CATEGORIAS:
        row = {}
        for cat, info in CATEGORIAS_PREDEFINIDAS.items():
            if info['grupo'] == grupo:
                row[cat] = df_categorizado[cat].sum()
        data[grupo] = row
    df_heatmap = pd.DataFrame(data).T
    fig = px.imshow(df_heatmap, text_auto=True, aspect="auto", color_continuous_scale="YlGnBu",
                    title="Mapa de calor: Distribución de categorías por grupo")
    return fig

def mostrar_datos_por_categoria(df_categorizado: pd.DataFrame):
    st.markdown("### Datos por categoría")
    
    # Organiza las categorías por grupo
    categorias_por_grupo = {}
    for cat, info in CATEGORIAS_PREDEFINIDAS.items():
        categorias_por_grupo.setdefault(info['grupo'], []).append(cat)
    
    for grupo, categorias in categorias_por_grupo.items():
        with st.expander(f"Grupo: {grupo}"):
            for cat in categorias:
                registros = df_categorizado[df_categorizado[cat] == 1]
                if not registros.empty:
                    st.markdown(f"**Categoría: {cat}** (Total: {len(registros)} registros)")
                    
                    # Crear una copia de los datos relevantes para mostrar con solo las columnas especificadas
                    columnas_deseadas = [
                        'PD_NOMBRECOMPLETOVICFORMULA', 
                        'PD_ESTATUSVIC', 
                        'PD_SEXO', 
                        'PD_EDAD', 
                        'DD_HECHOS', 
                        'DD_MPIO', 
                        'DD_ESTADO'
                    ]
                    
                    # Filtrar solo las columnas que existen en el dataframe
                    columnas_disponibles = [col for col in columnas_deseadas if col in registros.columns]
                    df_display = registros[columnas_disponibles].copy()
                    
                    # Limitar la longitud de DD_HECHOS para mejor visualización en la tabla
                    if 'DD_HECHOS' in df_display.columns:
                        df_display['DD_HECHOS'] = df_display['DD_HECHOS'].apply(
                            lambda x: x[:300] + "..." if len(str(x)) > 300 else x
                        )
                    
                    # Mostrar la tabla con todos los registros de esta categoría
                    st.dataframe(df_display)
                else:
                    st.markdown(f"**Categoría: {cat}** - No se encontraron registros.")
                
                # Añadir un separador entre categorías para mejor legibilidad
                if cat != categorias[-1]:
                    st.markdown("---")
    
    # Después de mostrar todas las categorías, mostrar la tabla de errores de edad
    df_errores = detectar_errores_edad(df_categorizado)
    with st.expander("Posibles errores de incongruencia en edades"):
        if not df_errores.empty:
            st.markdown(f"### Se encontraron {len(df_errores)} posibles errores de edad")
            st.dataframe(df_errores)
        else:
            st.markdown("No se encontraron incongruencias de edad.")

def detectar_errores_edad(df: pd.DataFrame) -> pd.DataFrame:
    """
    Detecta posibles incongruencias entre la edad y la categoría asignada.
    Mantiene los índices originales del DataFrame para identificar las filas.
    
    Args:
        df: DataFrame con los datos categorizados
        
    Returns:
        DataFrame con los posibles errores de incongruencia conservando índices originales
    """
    # Límites de edad por categoría
    limites_edad = {
        'CON_PAREJA': 10,
        'CON_AMIGOS': 10,
        'CONYUGE': 10, 
        'ABANDONO_HOGAR': 10,
        'FUGA_DE_ALBERGUE': 6,
        'ACTIVIDAD': 6,
        'CENTRO_REHABILITACION': 17,
        'DETENCION': 13,
        'SALIO_DE_CASA_NO_VOLVIO': 3,
        'DISCUSION': 4
    }
    
    # Columnas que queremos mostrar en la tabla de errores
    columnas_deseadas = [
        'PD_NOMBRECOMPLETOVICFORMULA', 
        'PD_ESTATUSVIC', 
        'PD_SEXO', 
        'PD_EDAD', 
        'DD_HECHOS', 
        'DD_MPIO', 
        'DD_ESTADO'
    ]
    
    # Diccionario para almacenar las filas con errores utilizando los índices originales como claves
    errores_por_indice = {}
    
    # Verificar cada categoría y sus límites de edad
    for categoria, limite_edad in limites_edad.items():
        if categoria in df.columns:
            # Filtrar los registros que pertenecen a la categoría
            registros_categoria = df[df[categoria] == 1]
            
            for idx, row in registros_categoria.iterrows():
                # Verificar si la edad existe y es un número
                if 'PD_EDAD' in row and pd.notna(row['PD_EDAD']):
                    try:
                        edad = float(row['PD_EDAD'])
                        # Verificar si hay incongruencia entre la edad y la categoría
                        if edad < limite_edad:
                            # Verificar si hay menciones específicas que podrían justificar la categoría
                            if 'DD_HECHOS' in row and pd.notna(row['DD_HECHOS']):
                                descripcion = str(row['DD_HECHOS']).lower()
                                # No considerar error si se menciona explícitamente hijo/niño
                                if not ('hij' in descripcion or 'niñ' in descripcion):
                                    # Guardar el índice original y la información de error
                                    if idx not in errores_por_indice:
                                        errores_por_indice[idx] = {
                                            col: row[col] for col in columnas_deseadas if col in row
                                        }
                                        # Agregar información sobre la categoría que causó el error
                                        errores_por_indice[idx]['Categoria'] = categoria
                                        errores_por_indice[idx]['Edad_Minima_Esperada'] = limite_edad
                                    else:
                                        # Si ya existe un error para este índice, actualizar la categoría
                                        errores_por_indice[idx]['Categoria'] += f", {categoria}"
                    except (ValueError, TypeError):
                        # Si la edad no se puede convertir a número, ignoramos esta fila
                        pass
    
    # Crear DataFrame con los errores encontrados
    if errores_por_indice:
        # Convertir diccionario a DataFrame conservando los índices originales
        df_errores = pd.DataFrame.from_dict(errores_por_indice, orient='index')
        
        # Limitar la longitud de DD_HECHOS para mejor visualización
        if 'DD_HECHOS' in df_errores.columns:
            df_errores['DD_HECHOS'] = df_errores['DD_HECHOS'].apply(
                lambda x: x[:300] + "..." if x and len(str(x)) > 300 else x
            )
        
        return df_errores
    else:
        return pd.DataFrame()

# ---------------------------
# 2da Parte: Nuevas Funciones para la pestaña "Análisis de Contexto"
# ---------------------------
@st.fragment
def plot_ultimo_lugar_trend(data: pd.DataFrame):
    """
    Grafica la evolución anual de los 7 últimos lugares más frecuentes donde se vio a la persona.
    Se muestra además una tabla y un heatmap con todos los lugares.
    """
    if "CI_FECDEN" not in data.columns or "DD_ULTLUGAR" not in data.columns:
        st.warning("No se encontraron las columnas necesarias ('CI_FECDEN' o 'DD_ULTLUGAR').")
        return None

    # Convertir fecha y extraer año
    data["AÑO_DENUNCIA"] = data["CI_FECDEN"].dt.year

    # Rango de años
    min_year = int(data["AÑO_DENUNCIA"].min())
    max_year = int(data["AÑO_DENUNCIA"].max())

    # Slider de años
    selected_range = st.slider(
        "Intervalo de años (Último Lugar)",
        min_year,
        max_year,
        (min_year, max_year),
        key="slider_ultlugar"
    )

    # Filtrar datos
    df_filtrado = data[(data["AÑO_DENUNCIA"] >= selected_range[0]) &
                       (data["AÑO_DENUNCIA"] <= selected_range[1])]

    # Top 7 lugares
    top_lugares = df_filtrado["DD_ULTLUGAR"].value_counts().head(7).index.tolist()

    # Agrupar para gráfico principal
    df_grouped = df_filtrado[df_filtrado["DD_ULTLUGAR"].isin(top_lugares)]
    df_grouped = df_grouped.groupby(["AÑO_DENUNCIA", "DD_ULTLUGAR"]).size().reset_index(name="Total")

    # Gráfico de líneas
    fig = px.line(
        df_grouped,
        x="AÑO_DENUNCIA",
        y="Total",
        color="DD_ULTLUGAR",
        markers=True,
        title="Tendencia de los últimos lugares más frecuentes donde fue vista la persona"
    )
    fig.update_layout(xaxis_title="Año", yaxis_title="Casos", legend_title="Último lugar")
    st.plotly_chart(fig, use_container_width=True)

    # Mostrar tabla resumen del top 7
    with st.expander("Ver datos en tabla"):
        st.dataframe(df_grouped.pivot(index="AÑO_DENUNCIA", columns="DD_ULTLUGAR", values="Total").fillna(0))

    # Tabla y heatmap de todos los lugares
    df_all = df_filtrado.groupby(["AÑO_DENUNCIA", "DD_ULTLUGAR"]).size().reset_index(name="Total")
    df_pivot = df_all.pivot(index="AÑO_DENUNCIA", columns="DD_ULTLUGAR", values="Total").fillna(0)

    with st.expander("Ver todos los lugares (tabla y mapa de calor)"):
        st.dataframe(df_pivot)

        # Top 19 lugares por total de casos
        top_20_lugares = df_all.groupby("DD_ULTLUGAR")["Total"].sum().nlargest(19).index.tolist()
        df_top20 = df_pivot[top_20_lugares]

                # Crear mapa de calor con escala de color personalizada (rangos fijos)
        fig_heatmap = px.imshow(
            df_top20.T,
            aspect="auto",
            zmin=0,
            zmax=650,  # valor máximo esperado para normalizar la escala
            color_continuous_scale=[
                [0.0, 'black'],     # 0
                [0.1, 'green'],     # ~65/650
                [0.3, 'yellow'],    # ~195/650
                [0.6, 'orange'],    # ~390/650
                [1.0, 'red']        # >=650
            ],
            labels={"x": "Año", "y": "Último Lugar", "color": "Casos"},
            title="Mapa de calor de los 19 lugares más frecuentes (con rangos de color definidos)"
        )



        col_grafico, col_leyenda = st.columns([6, 1])  # Relación 3:1

        with col_grafico:
            st.plotly_chart(fig_heatmap, use_container_width=True)

        with col_leyenda:
            # Agrega espacio para bajar la leyenda y alinearla visualmente
            st.markdown("<br><br><br>", unsafe_allow_html=True)
            st.markdown("##### Rangos de color:")
            st.markdown("""
            <div style='line-height: 2'>
            <span style="color:black">⬛</span> **0**<br>
            <span style="color:green">🟩</span> **1 – 64**<br>
            <span style="color:yellow">🟨</span> **65 – 194**<br>
            <span style="color:orange">🟧</span> **195 – 389**<br>
            <span style="color:red">🟥</span> **390 o más**
            </div>
            """, unsafe_allow_html=True)




@st.fragment
def plot_ocupacion_trend(data: pd.DataFrame):
    """
    Grafica la evolución anual de las 5 ocupaciones (oficio/profesión) más frecuentes.
    Se utiliza "CI_FECDEN" para filtrar por años y "PD_OCUPA" para agrupar.
    """
    if "CI_FECDEN" not in data.columns or "PD_OCUPA" not in data.columns:
        st.warning("No se encontraron las columnas necesarias ('CI_FECDEN' o 'PD_OCUPA').")
        return None

    año_denuncia = "AÑO_CI"

    min_year = int(data[año_denuncia].min())
    max_year = int(data[año_denuncia].max())
    selected_range = st.slider("Intervalo de años (Ocupación)", min_year, max_year, (min_year, max_year), key="slider_ocupa")
    df_filtrado = data[(data[año_denuncia] >= selected_range[0]) & (data[año_denuncia] <= selected_range[1])]

    # Identificar las 5 ocupaciones más frecuentes en el período seleccionado
    top_7 = df_filtrado["PD_OCUPA"].value_counts().head(7).index.tolist()

    # Agrupar por año y ocupación
    df_grouped = df_filtrado[df_filtrado["PD_OCUPA"].isin(top_7)]
    df_grouped = df_grouped.groupby([año_denuncia, "PD_OCUPA"]).size().reset_index(name="Total")

    fig = px.line(df_grouped, x=año_denuncia, y="Total", color="PD_OCUPA",
                  markers=True,
                  title="Tendencia de Ocupaciones de Personas Desaparecidas")
    
    fig.update_layout(xaxis_title="Año", legend_title="Ocupaciones")
    st.plotly_chart(fig, use_container_width=True)
    
    # Añadir un expander con la tabla de datos
    with st.expander("Ver datos de la gráfica"):
        # Ordenar datos para mejor visualización y Rellenar valores NaN con 0
        df_tabla = df_grouped.pivot(index=año_denuncia, columns="PD_OCUPA", values="Total").fillna(0)
        
        # Calcular los totales por columna
        totales = pd.DataFrame(df_tabla.sum(axis=0), columns=["Total"]).T

        # Añadir la fila de totales al DataFrame
        df_tabla_con_totales = pd.concat([df_tabla, totales])

        # Formatear la fila de totales para que la columna del año diga "Total"
        df_tabla_con_totales.rename(index={df_tabla_con_totales.index[-1]: "Total"}, inplace=True)

        # Mostrar la tabla
        st.dataframe(df_tabla_con_totales, use_container_width=False)


# -------------------------
# Tabla Registro Estatal (Denuncias)
# -------------------------
def generate_registro_estatal_table_denuncias(df: pd.DataFrame) -> pd.DataFrame:
    """
    Genera la tabla de Registro Estatal usando las fechas de denuncia (CI_FECDEN)
    con las columnas: AÑO, PERSONAS REPORTADAS, PERSONAS LOCALIZADAS, PENDIENTES.
    Se considera únicamente datos desde el año 2001 hasta el año actual y se agrega
    la fila TOTAL al final.
    """
    if "CI_FECDEN" not in df.columns or "PD_ESTATUSVIC" not in df.columns:
        st.warning("Columnas necesarias no encontradas para la tabla de denuncias.")
        return pd.DataFrame()
    
    df["AÑO_DENUNCIAS"] = df["CI_FECDEN"].dt.year
    df_den = df[df["AÑO_DENUNCIAS"] >= 2001]
    
    tabla = (df_den.groupby("AÑO_DENUNCIAS")
             .agg(PERSONAS_REPORTADAS=("AÑO_DENUNCIAS", "count"),
                  PERSONAS_LOCALIZADAS=("PD_ESTATUSVIC", lambda x: (x == "LOCALIZADA").sum()))
             .reset_index()
             .rename(columns={"AÑO_DENUNCIAS": "AÑO"}))
    
    # Convertir la columna AÑO a string para evitar problemas de conversión
    tabla["AÑO"] = tabla["AÑO"].astype(str)
    tabla["PENDIENTES"] = tabla["PERSONAS_REPORTADAS"] - tabla["PERSONAS_LOCALIZADAS"]
    
    totales = pd.DataFrame({
        "AÑO": ["TOTAL"],
        "PERSONAS_REPORTADAS": [tabla["PERSONAS_REPORTADAS"].sum()],
        "PERSONAS_LOCALIZADAS": [tabla["PERSONAS_LOCALIZADAS"].sum()],
        "PENDIENTES": [tabla["PENDIENTES"].sum()]
    })
    
    tabla = pd.concat([tabla, totales], ignore_index=True)
    return tabla


def style_registro_estatal(df: pd.DataFrame) -> str:
    """Devuelve la tabla estilizada en HTML sin índice."""
    styled = (df.style
              .apply(lambda row: ['background-color: #0000FF; color: white' if row['AÑO'] == 'TOTAL' else '' for _ in row],
                     axis=1)
              .set_table_styles([{'selector': 'th', 'props': [('background-color', 'lightblue'), ('color', 'black')]}])
              .hide(axis="index"))
    return styled.to_html()

def formato_fecha_actual() -> str:
    """Devuelve la fecha actual en el formato '08 de Febrero 2025'."""
    hoy = datetime.today()
    meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
             7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    return hoy.strftime("%d") + " de " + meses[hoy.month] + " " + hoy.strftime("%Y")

# -------------------------
# Shutdown_Signal
# -------------------------

# Ruta para el archivo de señal
signal_file = "shutdown_signal.txt"

# Función para verificar la señal
def check_shutdown_signal():
    while True:
        if os.path.exists(signal_file):
            # Elimina el archivo de señal
            os.remove(signal_file)
            print("Cerrando aplicación Streamlit...")
            # Envía señal SIGINT (equivalente a Ctrl+C)
            pid = os.getpid()
            os.kill(pid, signal.SIGINT)
            return
        time.sleep(1)


# -------------------------
# Funciones para cada pestaña
# -------------------------
def tab_desapariciones(data, data_complete):
    """Contenido de la pestaña Desapariciones."""
    st.header("No Localizados")
    
    total_desaparecidos, total_localizados = display_metrics(data)
    st.markdown("---")
    
    plot_desapariciones_por_año(data)
    
    st.write("---")
    # Fragmento de Función main():
    ejecutar_prediccion_arima(data_complete)
    #############################
    plot_registros_por_año(data)
    #plot_expedientes_por_año(data)
    
    st.markdown("---")
    #Registro Estatal (Denuncias)
    fecha_max_denuncia = data["CI_FECDEN"].max().date()
    tabla_registro = generate_registro_estatal_table_denuncias(data)
    if not tabla_registro.empty:
        st.subheader(f"Registro Estatal (Denuncias) corte al {fecha_max_denuncia} en fecha {formato_fecha_actual()}")
        st.markdown(style_registro_estatal(tabla_registro), unsafe_allow_html=True)

    
    st.write("---")
    validar_fechas_desaparicion(data)

def tab_reapariciones(data):
    """Contenido de la pestaña Reapariciones."""
    st.header("Localizaciones")
    display_metrics_reaparecidos(data)
    st.markdown("---")
    plot_reapariciones(data)
    plot_outliers_reapariciones(data)



import json
import unicodedata

@st.fragment
def generar_mapa_calor_desapariciones(df, ruta_geojson="Yucatan.json"):
    """
    Genera un mapa de calor interactivo de desapariciones por municipio en Yucatán.
    Los municipios se colorean según el número de casos en sus delimitaciones geopolíticas.

    Args:
        df (pd.DataFrame): DataFrame con las columnas "DD_MPIO" (municipio),
                           "PD_ESTATUSVIC" (estatus de la investigación) y "AÑO_CI" (año del reporte).
        ruta_geojson (str): Ruta al archivo GeoJSON con las delimitaciones de los municipios.

    Returns:
        streamlit.plotly_chart: Un objeto de gráfico de Plotly para Streamlit.
    """

    
    def normalizar_nombre_municipio(nombre):
        """
        Normaliza nombres de municipios removiendo acentos y convirtiendo a mayúsculas.
        Ej: "Sinanché" -> "SINANCHE"
        """
        if pd.isna(nombre) or nombre is None:
            return ""
        
        # Convertir a string y hacer strip
        nombre_str = str(nombre).strip()
        
        # Remover acentos usando unicodedata
        nombre_sin_acentos = unicodedata.normalize('NFD', nombre_str)
        nombre_sin_acentos = ''.join(char for char in nombre_sin_acentos 
                                   if unicodedata.category(char) != 'Mn')
        
        # Convertir a mayúsculas
        return nombre_sin_acentos.upper()

    # Verificar columnas necesarias
    if "DD_MPIO" not in df.columns or "PD_ESTATUSVIC" not in df.columns or "AÑO_CI" not in df.columns:
        st.warning("No se encontraron las columnas necesarias ('DD_MPIO', 'PD_ESTATUSVIC' o 'AÑO_CI').")
        return None

    # Cargar datos geoespaciales
    try:
        with open(ruta_geojson, 'r', encoding='utf-8') as f:
            geojson_data = json.load(f)
        st.success(f"✅ Archivo GeoJSON cargado correctamente: {len(geojson_data['features'])} municipios encontrados")
    except FileNotFoundError:
        st.error(f"❌ No se encontró el archivo GeoJSON: {ruta_geojson}")
        st.info("💡 Necesitas un archivo GeoJSON con las delimitaciones de los municipios de Yucatán.")
        st.info("📁 Puedes descargarlo de fuentes como INEGI o portales de datos abiertos del gobierno.")
        return None
    except json.JSONDecodeError:
        st.error("❌ El archivo GeoJSON no tiene un formato válido.")
        return None

    # Controles de interfaz mejorados
    st.markdown("### 🗺️ Mapa de Calor - Personas Desaparecidas en Yucatán")
    
    año_reporte = "AÑO_CI"
    min_year = int(df[año_reporte].min())
    max_year = int(df[año_reporte].max())
    
    col1, col2 = st.columns([3, 1])
    with col1:
        selected_range = st.slider(
            "📅 Selecciona el período de análisis", 
            min_year, 
            max_year, 
            (min_year, max_year), 
            key="slider_mapa_calor",
            help="Ajusta el rango para ver los casos reportados en diferentes períodos"
        )
    
    with col2:
        st.metric(
            "Período seleccionado", 
            f"{selected_range[1] - selected_range[0] + 1} años",
            f"{selected_range[0]} - {selected_range[1]}"
        )
    
    # Filtrar por año
    df_filtrado_año = df[
        (df[año_reporte] >= selected_range[0]) & 
        (df[año_reporte] <= selected_range[1])
    ].copy()

    # Filtrar solo los casos "EN INVESTIGACION" para el mapa
    df_investigacion = df_filtrado_año[
        df_filtrado_año['PD_ESTATUSVIC'] == 'EN INVESTIGACION'
    ].copy()

    # Normalizar nombres de municipios
    df_investigacion['DD_MPIO_NORM'] = df_investigacion['DD_MPIO'].str.strip().str.upper()

    # Agrupar por municipio y contar los casos
    conteo_municipios = df_investigacion['DD_MPIO_NORM'].value_counts().reset_index()
    conteo_municipios.columns = ['municipio', 'num_casos']

    # Crear diccionario de casos por municipio para facilitar el mapeo
    casos_dict = dict(zip(conteo_municipios['municipio'], conteo_municipios['num_casos']))

    # Extraer nombres de municipios del GeoJSON y mapear casos
    municipios_geojson = []
    casos_mapeados = []
    
    for feature in geojson_data['features']:
        # Tu archivo Yucatan.json usa 'NOMGEO' para los nombres de municipios
        nombre_mpio = feature['properties'].get('NOMGEO')
        
        if nombre_mpio:
            # Normalizar el nombre del JSON (quitar acentos y convertir a mayúsculas)
            nombre_normalizado = normalizar_nombre_municipio(nombre_mpio)
            municipios_geojson.append(nombre_normalizado)
            casos_mapeados.append(casos_dict.get(nombre_normalizado, 0))
            
            # Guardar también el nombre original para mostrar en el mapa
            feature['properties']['NOMGEO_NORM'] = nombre_normalizado
        else:
            st.warning(f"⚠️ Feature sin nombre de municipio encontrada en el JSON")

    # Crear DataFrame para el mapa
    df_mapa = pd.DataFrame({
        'municipio': municipios_geojson,
        'num_casos': casos_mapeados
    })

    # Verificar municipios sin datos
    municipios_sin_casos = df_mapa[df_mapa['num_casos'] == 0]['municipio'].tolist()
    municipios_con_casos = df_mapa[df_mapa['num_casos'] > 0]
    
    if municipios_sin_casos and len(municipios_sin_casos) < 20:  # Solo mostrar si no son demasiados
        st.info(f"ℹ️ Municipios sin casos registrados: {', '.join(municipios_sin_casos[:10])}" + 
                (f" y {len(municipios_sin_casos)-10} más..." if len(municipios_sin_casos) > 10 else ""))

    # Verificar si hay municipios en los datos que no están en el GeoJSON
    municipios_datos = set(conteo_municipios['municipio'].tolist())
    municipios_geo = set(municipios_geojson)
    municipios_faltantes = municipios_datos - municipios_geo
    
    if municipios_faltantes:
        st.warning(f"⚠️ Municipios en los datos pero no en el MAPA: {', '.join(list(municipios_faltantes)[:5])}" +
                  (f" y {len(municipios_faltantes)-5} más..." if len(municipios_faltantes) > 5 else ""))

    # Crear categorías de riesgo
    if not df_mapa.empty and df_mapa['num_casos'].max() > 0:
        df_mapa['categoria_riesgo'] = pd.cut(
            df_mapa['num_casos'], 
            bins=[0, 1, 10, 25, 50, float('inf')], 
            labels=['Sin casos', 'Bajo (1-10)', 'Medio (11-25)', 'Alto (26-50)', 'Muy Alto (50+)'],
            include_lowest=True
        )
        
        # Crear el mapa de calor con choropleth
        fig = px.choropleth_mapbox(
            df_mapa,
            geojson=geojson_data,
            locations='municipio',
            featureidkey="properties.NOMGEO_NORM",  # Usamos el nombre normalizado para el match
            color='num_casos',
            hover_name='municipio',
            hover_data={'num_casos': True},
            mapbox_style="open-street-map",
            zoom=7,
            center={"lat": 20.7, "lon": -89.1},  # Centro de Yucatán
            opacity=0.7,
            color_continuous_scale=[
                [0, "#F8F9FA"],      # Blanco para 0 casos
                #[0.2, "#FFF3CD"],    # Amarillo claro
                [0.03, "#FFE53B"],    # Amarillo
                [0.08, "#C59E00"],    # Amarillo fuerte
                [0.2, "#FF8C00"],    # Naranja
                [0.5, "#DC3545"] ,    # Rojo
                [1.0, "#980A18"]     # Rojo
            ],
            title=f"🗺️ Mapa de Calor - Casos de Personas Desaparecidas (En Investigación)<br>Yucatán • {selected_range[0]}-{selected_range[1]}",
            labels={'num_casos': 'Casos en Investigación'}
        )
        
        # Personalizar el hover
        fig.update_traces(
            hovertemplate="<b>%{hovertext}</b><br>" +
                         "📊 Casos en investigación: %{z}<br>" +
                         "<extra></extra>",
            marker_line_width=1,
            marker_line_color='black'
        )
        
        # Mejorar el layout
        fig.update_layout(
            height=700,
            margin={"r":10,"t":100,"l":10,"b":10},
            title={
                'text': f"🗺️ Mapa de Calor - Casos de Personas Desaparecidas (En Investigación)<br><sub>Yucatán • {selected_range[0]}-{selected_range[1]} • Total: {df_mapa['num_casos'].sum()} casos en {len(municipios_con_casos)} municipios</sub>",
                'x': 0.5,
                'xanchor': 'center',
                'font': {'size': 16, 'color': "#006EFF"}
            },
            coloraxis_colorbar=dict(
                title="Casos en<br>Investigación",
                thickness=20,
                len=0.7,
                x=1.02,
                tickmode="linear",
                tick0=0,
                dtick=max(1, df_mapa['num_casos'].max() // 5) if df_mapa['num_casos'].max() > 0 else 1
            )
        )
        
        st.plotly_chart(fig, use_container_width=True)

        # Mostrar estadísticas mejoradas
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "📍 Municipios afectados", 
                len(municipios_con_casos),
                f"de {len(df_mapa)} total",
                help="Número de municipios con casos en investigación"
            )
        
        with col2:
            st.metric(
                "📊 Total de casos", 
                df_mapa['num_casos'].sum(),
                help="Suma total de casos en investigación"
            )
        
        with col3:
            municipio_max = df_mapa.loc[df_mapa['num_casos'].idxmax()]
            st.metric(
                "📈 Casos máximos", 
                int(municipio_max['num_casos']),
                f"en {municipio_max['municipio']}",
                help="Municipio con mayor número de casos"
            )
        
        with col4:
            promedio = df_mapa[df_mapa['num_casos'] > 0]['num_casos'].mean() if len(municipios_con_casos) > 0 else 0
            st.metric(
                "📊 Promedio por municipio", 
                f"{promedio:.1f}",
                help="Promedio de casos por municipio afectado"
            )

        # Mostrar distribución por categorías
        # Función para categorizar dinámicamente basado en porcentajes
        def categorizar_riesgo_por_porcentajes(casos, max_casos):
            """
            Categoriza el riesgo basándose en porcentajes del valor máximo actual
            """
            if casos == 0:
                return 'Sin casos'
            
            # Calcular el porcentaje respecto al máximo
            porcentaje = (casos / max_casos) * 100
            
            # Dividir en 4 categorías para casos > 0 (20% cada una)
            if porcentaje <= 20:
                return 'Bajo'
            elif porcentaje <= 40:
                return 'Medio'
            elif porcentaje <= 70:
                return 'Alto'
            else:
                return 'Muy Alto'

        # Mostrar distribución por categorías
        st.markdown("### 📊 Distribución por Categorías de Riesgo")

        # Calcular el valor máximo actual para hacer la categorización adaptativa
        max_casos_actual = df_mapa['num_casos'].max() if len(df_mapa) > 0 else 0

        # Solo proceder si hay datos
        if max_casos_actual > 0:
            # Recategorizar dinámicamente basado en porcentajes
            df_mapa['categoria_riesgo_adaptativa'] = df_mapa['num_casos'].apply(
                lambda x: categorizar_riesgo_por_porcentajes(x, max_casos_actual)
            )
            
            # Definir el orden de las categorías
            orden_categorias = ['Sin casos', 'Bajo', 'Medio', 'Alto', 'Muy Alto']
            
            # Contar categorías y mantener el orden
            categoria_counts = df_mapa['categoria_riesgo_adaptativa'].value_counts().reindex(orden_categorias, fill_value=0)
            # Remover categorías con 0 casos para la visualización
            categoria_counts = categoria_counts[categoria_counts > 0]
            
            # Definir colores y emojis fijos para cada categoría
            colores_por_categoria = {
                'Sin casos': "#F8F9FA",    # Blanco
                'Bajo': "#FFEC74",         # Amarillo claro
                'Medio': "#FFB700",        # Amarillo oscuro
                'Alto': "#FF7B00",         # Naranja
                'Muy Alto': "#DC3545"      # Rojo
            }
            
            emojis_por_categoria = {
                'Sin casos': "⚪",     # Blanco
                'Bajo': "🟡",          # Amarillo claro
                'Medio': "🟨",         # Amarillo oscuro
                'Alto': "🟠",          # Naranja
                'Muy Alto': "🔴"       # Rojo
            }
            
            # Obtener colores en el orden correcto
            colores_ordenados = [colores_por_categoria[cat] for cat in categoria_counts.index]
            
            # Calcular rangos dinámicos para mostrar en la leyenda
            def obtener_rangos_dinamicos(max_casos):
                if max_casos == 0:
                    return {}
                
                rangos = {}
                rangos['Sin casos'] = "0 casos/denuncias"
                rangos['Bajo'] = f"1 - {int(max_casos * 0.2)} casos/denuncias"
                rangos['Medio'] = f"{int(max_casos * 0.2) + 1} - {int(max_casos * 0.4)} casos/denuncias"
                rangos['Alto'] = f"{int(max_casos * 0.4) + 1} - {int(max_casos * 0.7)} casos/denuncias"
                rangos['Muy Alto'] = f"{int(max_casos * 0.7) + 1} - {max_casos} casos/denuncias"
                
                return rangos
            
            rangos_dinamicos = obtener_rangos_dinamicos(max_casos_actual)
            
            col1, col2 = st.columns([2, 1])
            with col1:
                fig_bar = px.bar(
                    x=categoria_counts.index,
                    y=categoria_counts.values,
                    title=f"Municipios por Categoría de Riesgo (Máx: {max_casos_actual} casos)",
                    labels={'x': 'Categoría', 'y': 'Número de Municipios'},
                    color=categoria_counts.index,
                    color_discrete_sequence=colores_ordenados
                )
                fig_bar.update_layout(height=450, showlegend=False)
                
                # Agregar anotaciones con el número de municipios encima de cada barra
                fig_bar.update_traces(
                    texttemplate='%{y}',
                    textposition='outside'
                )
                
                st.plotly_chart(fig_bar, use_container_width=True)
            
            with col2:
                st.markdown("#### 🎯 Leyenda de Categorías")
                st.markdown(f"*Rangos basados en porcentajes del máximo ({max_casos_actual} casos)*")
                
                for categoria, count in categoria_counts.items():
                    emoji = emojis_por_categoria.get(categoria, "⚫")
                    porcentaje = (count / len(df_mapa)) * 100
                    rango = rangos_dinamicos.get(categoria, "")
                    st.markdown(f"{emoji} **{categoria}**: {count} municipios ({porcentaje:.1f}%)")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;↳ {rango}")

        else:
            st.warning("No hay casos registrados en el período seleccionado.")

        # Mostrar tabla de casos por municipio con mejores estilos
        # Define tus colores personalizados
        colores_personalizados = ["#FFE53B", "#C59E00", "#FF8C00", "#DC3545"]

        # Crea un mapa de colores personalizado
        from matplotlib.colors import LinearSegmentedColormap
        cmap_personalizado = LinearSegmentedColormap.from_list(
            "mi_gradiente_personalizado", colores_personalizados, N=256
        )

        with st.expander("📋 Ver detalle de casos por municipio"):            
            # Ordenar por número de casos descendente
            display_df = df_mapa[df_mapa['num_casos'] > 0].sort_values('num_casos', ascending=False)
            
            if len(display_df) > 0:
                # Usar la categorización adaptativa también en la tabla
                display_df_formatted = display_df[['municipio', 'num_casos', 'categoria_riesgo_adaptativa']].rename(columns={
                    'municipio': '🏛️ Municipio',
                    'num_casos': '📊 Casos en Investigación',
                    'categoria_riesgo_adaptativa': '🎯 Categoría de Riesgo'
                })
                
                st.dataframe(
                    display_df_formatted.style.background_gradient(
                        subset=['📊 Casos en Investigación'], 
                        cmap=cmap_personalizado
                    ),
                    use_container_width=True
                )
                
                # Mostrar estadísticas adicionales
                st.markdown("##### 📈 Estadísticas del período seleccionado:")
                col_stats1, col_stats2, col_stats3 = st.columns(3)
                
                col_stats1.metric("Total de denuncias", display_df['num_casos'].sum())
                col_stats2.metric("Municipios afectados", len(display_df))
                col_stats3.metric("Promedio por municipio", f"{display_df['num_casos'].mean():.1f}")
                    
                # Mostrar distribución de casos por categoría
                st.markdown("##### 📊 Distribución de casos por categoría:")
                casos_por_categoria = display_df.groupby('categoria_riesgo_adaptativa')['num_casos'].sum().sort_index()
                
                for categoria, total_casos in casos_por_categoria.items():
                    emoji = emojis_por_categoria.get(categoria, "⚫")
                    porcentaje_casos = (total_casos / display_df['num_casos'].sum()) * 100
                    st.markdown(f"{emoji} **{categoria}**: {total_casos} denuncias ({porcentaje_casos:.1f}% del total)")
                
            else:
                st.info("No hay casos registrados en el período seleccionado.")

    else:
        st.error("❌ No se encontraron datos para generar el mapa de calor.")
        st.info("💡 Verifica que:")
        st.write("- El archivo GeoJSON contiene las delimitaciones correctas")
        st.write("- Los nombres de municipios coinciden entre los datos y el GeoJSON")
        st.write("- Hay casos 'EN INVESTIGACION' en el período seleccionado")

#Ya ha sido validado, ya no es necesario,
def cargar_geojson_yucatan(ruta_geojson):
    """
    Función auxiliar para cargar y validar el archivo JSON de municipios de Yucatán.
    
    Args:
        ruta_geojson (str): Ruta al archivo JSON (ej: "Yucatan.json")
        
    Returns:
        dict: Datos GeoJSON o None si hay error
    """
    try:
        with open(ruta_geojson, 'r', encoding='utf-8') as f:
            geojson_data = json.load(f)
        
        # Validar estructura básica
        if 'features' not in geojson_data:
            st.error("❌ El archivo JSON no tiene la estructura correcta (falta 'features')")
            return None
            
        if not geojson_data['features']:
            st.error("❌ El archivo JSON no contiene ninguna feature")
            return None
            
        return geojson_data
        
    except FileNotFoundError:
        st.error(f"❌ No se encontró el archivo JSON: {ruta_geojson}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"❌ Error al leer el archivo JSON: {str(e)}")
        return None
    except Exception as e:
        st.error(f"❌ Error inesperado al cargar el archivo JSON: {str(e)}")
        return None


def tab_demograficos(data):
    """Contenido de la pestaña Demográficos."""
    st.header("Demográficos")
    df_ejemplo = data
    generar_mapa_calor_desapariciones(df_ejemplo)

    
    plot_sexo_distribution(data)
    plot_clasedad_distribution(data)
    plot_age_distribution(data)
    st.markdown("---")
    plot_municipio_estados_distribution(data)
    st.markdown("---")

@st.fragment
def analisis_contexto_1parte(data):

    # Filtrar por año si existe la columna de fecha de desaparición
    if "CI_FECDEN" in data.columns:
        data["AÑO_DENUNCIA"] = data["CI_FECDEN"].dt.year
        min_year = int(data["AÑO_DENUNCIA"].min())
        max_year = int(data["AÑO_DENUNCIA"].max())
        selected_range = st.slider("Analisis Contexto (Denuncias)", min_year, max_year, (min_year, max_year))
        data_filtered = data[(data["AÑO_DENUNCIA"] >= selected_range[0]) & 
                             (data["AÑO_DENUNCIA"] <= selected_range[1])]
    else:
        st.warning("La columna 'CI_FECDEN' no se encontró. Se usarán todos los datos.")
        data_filtered = data

    # Procesar y categorizar registros
    df_categorizado = categorizar_registros(data_filtered)
    if df_categorizado is None:
        st.error("No se pudieron categorizar los registros.")
        return
    
    st.subheader("Resumen de Categorías")
    st.plotly_chart(plot_resumen_categorias(df_categorizado), use_container_width=True)

    st.subheader("Comparativa de Grupos")
    st.plotly_chart(plot_comparativa_grupos(df_categorizado), use_container_width=True)

    st.subheader("Heatmap de Categorías por Grupo")
    st.plotly_chart(plot_heatmap_categorias_por_grupo(df_categorizado), use_container_width=True)

    mostrar_datos_por_categoria(df_categorizado)

def tab_analisis_contexto(data: pd.DataFrame):
    # Procesar y categorizar registros
    df_categorizado = categorizar_registros(data)
    if df_categorizado is None:
        st.error("No se pudieron categorizar los registros.")
        return
    
    st.header("Análisis de Categorías y Contexto")
    display_metrics_contexto(df_categorizado)
    analisis_contexto_1parte(data)

    # Nuevas gráficas de tendencia
    st.markdown("---")
    st.subheader("Tendencia de Último Lugar Visto")
    plot_ultimo_lugar_trend(data)

    st.markdown("---")
    st.subheader("Tendencia de Ocupaciones de Personas Desaparecidas")
    plot_ocupacion_trend(data)


# -------------------------
# Función principal
# -------------------------
def main():
    """Función principal de la aplicación."""
    setup_page_config()
    st.title("📊 :blue[Monitoreo de Personas Desaparecidas]")

    # Cargar y limpiar datos (con caché)
    default_filepath = r"C:\Users\Angel\Documents\Programacion\Fiscalia\data_desaparecidos.xlsx"

    ###################################
    ## ------   SUBIR EXCEL   ------ ##
    
    # Opción de subir archivo
    st.sidebar.markdown("<h1 style='text-align: center;'>Carga de Datos</h1>", unsafe_allow_html=True)
    uploaded_file = st.sidebar.file_uploader(":violet[Subir archivo Excel]", type=['xlsx', 'xls'])
    
    # Definir las columnas que necesitas
    columnas_necesarias = ["AÑO_CI", "CI_CARPEINV", "PD_NOMBRECOMPLETOVICFORMULA", "PD_ESTATUSVIC", "PD_EDAD", "PD_SEXO", "PD_CLASEDAD", 
                           "PD_OCUPA", "DD_FECDESAP", "DD_FECPERCA", "CI_FECDEN", "DL_FECLOC", "DD_HECHOS", "DD_HIPOTESIS","DD_ULTLUGAR", 
                           "DD_LOCALIDAD", "DD_MPIO", "DD_ESTADO", "DL_LOCALSVCV", "DL_LOCALIDAD", "DL_MPIO", "DL_ESTADO"]

    # Cargar datos (con caché para optimizar)
    data, archivo_actual = load_data(
        uploaded_file=uploaded_file, 
        default_filepath=default_filepath,
        columns_to_use=columnas_necesarias
    )

    # Mostrar información del dataset en la barra lateral
    if archivo_actual:
        st.sidebar.info(f"Archivo actual: {archivo_actual}")
    ###################################
    
    if data is None: return
    data = clean_data(data)
    data_complete = data.copy()
    
    
    # Aplicar filtros en la barra lateral
    st.sidebar.markdown("<h1 style='text-align: center;'>Filtros</h1>", unsafe_allow_html=True)
    # Inicializar variables de estado para controlar los widgets
    if 'filter_key' not in st.session_state:
        st.session_state.filter_key = 0
    data_filtered = create_date_filter(data)
    

    ####################################################
    # Aplicar filtros de selección múltiple

    # Crear un key único para cada widget basado en filter_key
    current_key = st.session_state.filter_key

    # Filtro por estatus (denuncias) con key dinámico
    estados = st.sidebar.multiselect("Seleccione el :violet[Estatus] de las Personas",
                                      options=data_complete["PD_ESTATUSVIC"].unique(),
                                      default=data_complete["PD_ESTATUSVIC"].unique().tolist(),
                                      key=f"estados_{current_key}")
    data_filtered = data_filtered[data_filtered["PD_ESTATUSVIC"].isin(estados)]

    # Filtro por sexo (denuncias) con key dinámico
    sexo = st.sidebar.multiselect("Seleccione el :violet[Sexo] de las Personas",
                                      options=data_complete["PD_SEXO"].unique(),
                                      default=data_complete["PD_SEXO"].unique().tolist(),
                                      key=f"sexo_{current_key}")
    data_filtered = data_filtered[data_filtered["PD_SEXO"].isin(sexo)]

    # Filtro por clase de edad (denuncias) con key dinámico
    clases_edad = st.sidebar.multiselect("Seleccione la :violet[Clase] de Edad de Personas",
                                      options=data_complete["PD_CLASEDAD"].unique(),
                                      default=data_complete["PD_CLASEDAD"].unique().tolist(),
                                      key=f"edad_{current_key}")
    data_filtered = data_filtered[data_filtered["PD_CLASEDAD"].isin(clases_edad)]
    
    if data_filtered.empty:
        st.error("No hay datos para los filtros seleccionados.")
        return
    
    ###################################

    # Inicia el hilo de verificación (una sola vez)
    if "shutdown_thread_started" not in st.session_state:
        shutdown_thread = threading.Thread(target=check_shutdown_signal)
        shutdown_thread.daemon = True
        shutdown_thread.start()
        st.session_state.shutdown_thread_started = True
    
    # Inicializar estados si no existen
    if "show_confirm" not in st.session_state:
        st.session_state.show_confirm = False
    
    # Botón para mostrar confirmación
    if st.sidebar.button("Cerrar Aplicación"):
        st.session_state.show_confirm = True
    
    # Mostrar confirmación si el estado lo requiere
    if st.session_state.show_confirm:
        confirm = st.sidebar.checkbox("¿Estás seguro?")
        if confirm:
            with open(signal_file, "w") as f:
                f.write("shutdown")
            st.sidebar.success("Cerrando la aplicación...")
            st.stop()  # Detiene la ejecución del script actual
    ###################################
    # Crear pestañas
    tabs = st.tabs(["No Localizados", "Localizados", "Demográficos", "Análisis de Contexto"])
    
    # Contenido de cada pestaña
    with tabs[0]:
        tab_desapariciones(data_filtered, data_complete)
    
    with tabs[1]:
        tab_reapariciones(data_filtered)
    
    with tabs[2]:
        tab_demograficos(data_filtered)
    
    with tabs[3]:
        tab_analisis_contexto(data_filtered)

if __name__ == '__main__':
    main()